# -*- coding: utf-8 -*-
"""
合并工具（含 客户发言数 + 客服发言数 + 历史重点提升 + 标红逻辑）

本版要点：
1) 自动识别列：联系人 / 会话id / 会话消息内容 / 访客标签 / 接待成员
2) 根据聊天内容计算：
   - 客户发言综合标记：0 / 1 / "重点"
   - 客户发言数
   - 客服发言数
3) 期望访客标签（澳）：0互动 / 一般客户 / 重点 / 低意向客户（列名保留，内容换成宽泛类别）
4) 标红逻辑：
   - 用【客户发言综合标记】 vs 访客标签中的关键词 对比：
       · 标签含 0互动       -> 对应 0
       · 标签含 重点/全款/定金/分期 -> 对应 "重点"
       · 标签含 一般        -> 对应 1
       · 其它情况           -> 不参与判错（不标红）
   - 成交客户（全款 / 定金 / 分期）无论如何都不标红
   - 如果某个联系人的访客标签中含 whatsapp，则该联系人所有记录的 0/1/重点 按访客标签来统一决定
5) 时间类访客标签：
   - 支持 11.20 / 11/20 / 11月20 / 11-20 这几种格式
   - 用于历史重点提升判断
6) “按接待成员统计”中，仍然按记录数统计 0互动/一般/重点/低意向、按联系人统计成交和历史重点提升

新增规则（本次重点）：
7) 若某联系人：
   - 列【是否成交】为 True（访客标签中含 全款/定金/分期）
   - 且拥有 3 个及以上时间类访客标签
   => 仅统计为【成交】，不进入【0互动 / 一般 / 重点 / 历史重点提升】。
   且对该联系人跳过 whatsapp 覆盖逻辑，防止 whatsapp 把其强行改成重点或一般。
"""
import os
import re
import sys
import zipfile
import signal
import pandas as pd
from datetime import datetime
from tkinter import Tk, messagebox, Toplevel, Frame
import tkinter.filedialog as filedialog
from tkinter import ttk
import tkinter as tk

# ========= 可选：拖放支持（需要 tkinterdnd2） =========
try:
    from tkinterdnd2 import DND_FILES, TkinterDnD  # type: ignore
    _DND_AVAILABLE = True
except Exception:
    DND_FILES = None  # type: ignore
    TkinterDnD = None  # type: ignore
    _DND_AVAILABLE = False
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import Workbook

# -------- 兼容性补丁 --------
try:
    import pandas.io.formats.printing as _pd_printing
    if not hasattr(_pd_printing, "TerminalWriter"):
        class _DummyWriter:
            def __init__(self, *a, **k): pass
            def write(self, *a, **k): pass
        _pd_printing.TerminalWriter = _DummyWriter  # type: ignore
except Exception:
    pass

# ===== 全局变量初始声明（方便编辑器类型检查） =====
deal_three_labels_contacts = set()
contact_col = session_col = message_col = visitor_tag_col = agent_col = None

# ======================= 工具函数 =======================

def app_dir() -> str:
    if getattr(sys, 'frozen', False) and hasattr(sys, 'executable'):
        return os.path.dirname(os.path.abspath(sys.executable))
    return os.path.dirname(os.path.abspath(__file__))

def find_column(cols, candidates, contains=False):
    norm = lambda s: s.replace(' ', '').lower()
    norm_cols = {col: norm(col) for col in cols}
    if not contains:
        for cand in candidates:
            n_cand = norm(cand)
            for col, ncol in norm_cols.items():
                if ncol == n_cand:
                    return col
    for cand in candidates:
        n_cand = norm(cand)
        for col, ncol in norm_cols.items():
            if n_cand in ncol:
                return col
    return None

def first_non_null(series):
    s = series.dropna()
    return s.iloc[0] if not s.empty else None

# ======== 文本解析：提取"客户 : …"与"客服 : …"的所有片段 ========
CUSTOMER_RE = re.compile(r'客户\s*:\s*[“"]?\s*(.*?)[”"]?(?=$|\r|\n)')
AGENT_RE    = re.compile(r'客服\s*:\s*[“"]?\s*(.*?)[”"]?(?=$|\r|\n)')

IMG_TOKEN = "图片"
EMOJI_POWER_RE = re.compile(r'^【\s*(?:💪\s*)?\d+\+\s*】$')

def extract_customer_segments(text):
    if not isinstance(text, str):
        return []
    return [(seg or '').strip() for seg in CUSTOMER_RE.findall(text)]

def extract_agent_segments(text):
    if not isinstance(text, str):
        return []
    return [(seg or '').strip() for seg in AGENT_RE.findall(text)]


def is_zero_by_customer(segments):
    """
    判断是否“0互动”（最终稳定版）：
    1. 第二条发言是 "1" -> 一般
    2. 只有一次无意义行为 -> 0互动
    3. 出现有效文本 -> 一般
    """
    if not segments:
        return True

    segs = [(seg or "").strip() for seg in segments if isinstance(seg, str) and (seg or "").strip() != ""]

    # ⭐ 规则一：第二条是 1 -> 一般
    if len(segs) >= 2 and segs[1] == "1":
        return False

    ones_count = sum(1 for s in segs if s == "1")

    def _is_invalid(s: str) -> bool:
        return (s == "1") or (s == IMG_TOKEN) or bool(EMOJI_POWER_RE.match(s))

    # 规则二：出现任何有效文本 -> 一般
    if any(not _is_invalid(s) for s in segs):
        return False

    # 规则三：全是无效内容
    if ones_count >= 2:
        return False

    return True


def count_valid_customer_segments(text: str) -> int:
    """
    客户发言数：统计 '客户 :' 片段里，非图片/非表情/非孤立“1”的条数
    """
    if not isinstance(text, str):
        return 0
    segs = extract_customer_segments(text)
    return sum(1 for seg in segs if seg != IMG_TOKEN and not EMOJI_POWER_RE.match(seg) and seg != "1")

def count_valid_agent_segments(text: str) -> int:
    """
    客服发言数：统计 '客服 :' 片段里，
    只排除纯表情【💪40+】和孤立的 “1”，
    图片也算作一次有效发言。
    """
    if not isinstance(text, str):
        return 0
    segs = extract_agent_segments(text)
    # 不再排除 IMG_TOKEN，让“图片”也计入客服发言数
    return sum(
        1
        for seg in segs
        if not EMOJI_POWER_RE.match(seg) and seg != "1"
    )

def agent_focus_hit(text):
    """
    判断客服话术里是否出现重点话术/重点关键词，用于综合标记“重点”
    """
    parts = extract_agent_segments(text)
    if not parts:
        return False
    keywords = [
        "Highly effective concentration program",
        "Economic Concentration Program",
        "Experience Concentration Programme",
        "Concentration Programme",
        "High Concentration Program",
        "Economy Concentration Program",
        "Concentration Program",
        "Effective Concentration Experience Program",
        "Concentration Experience Program",
        "Experience Program",
        "concentration program",
        "Ένα εξαιρετικά αποτελεσματικό πρόγραμμα συγκέντρωσης",
        "Πρόγραμμα οικονομικής συγκέντρωσης",
        "οικονομικής συγκέντρωσης",
        "Programa de concentração económica",
        "Um programa de concentração de alta eficiência",
        "efficient concentration solution",
        "Economic Concentrate Formula",
        "Concentrate Formula",
        "برنامج تركيز عالي الكفاءة",
        "ب برنامج التركيز الاقتصادي",
        "Concentration Formula",
        "Strength Formula",
        "Uma fórmula de alta potência e concentração",
        "Fórmula para a força económica",
        "potência e concentração",
        "Concentrated Formula",
        "High-Efficiency Concentration Plan",
        "Economic Concentration Plan",
        "Concentration Plan",
        "Efficiency Concentration Plan",
        "efficiency concentration scheme",
        "concentration scheme",
        "Economic Concentration scheme"
        

    ]
    kw_lower = [k.lower() for k in keywords]
    for seg in parts:
        seg_low = seg.lower()
        if any(k in seg_low for k in kw_lower):
            return True
    return False

def combined_marker(text):
    """
    客户发言综合标记：
    - 如果客服重点话术命中 -> "重点"
    - 否则，根据客户发言是否为 0互动 -> 0 / 1
    """
    if agent_focus_hit(text):
        return "重点"
    cust = extract_customer_segments(text)
    if is_zero_by_customer(cust):
        return 0
    else:
        return 1

def read_any_excel(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsx":
        if not zipfile.is_zipfile(path):
            raise ValueError(f"文件不是合法的 .xlsx（zip）：{path}")
        return pd.read_excel(path, engine="openpyxl")
    elif ext == ".xls":
        try:
            import xlrd  # type: ignore
            return pd.read_excel(path, engine="xlrd")
        except Exception as e:
            raise RuntimeError("读取 .xls 失败。请安装 xlrd==1.2.0") from e
    else:
        raise ValueError(f"不支持的扩展名：{ext}（仅支持 .xlsx/.xls）")

def update_agent_re(agent_names):
    """
    动态构建客服（坐席）发言匹配正则，例如：
    (?:客服|王栋栋|黄经理) : “……”
    """
    global AGENT_RE

    speaker_pattern = r'(?:客服|' + "|".join(re.escape(n) for n in agent_names) + r')'
    pattern = speaker_pattern + r'\s*:\s*[“"]?\s*(.*?)[”"]?(?=$|\r|\n)'
    AGENT_RE = re.compile(pattern)


# ========= 时间标签识别 =========
# 支持：11.20 / 11/20 / 11月20 / 11-20 / 11..20 / 11,,20 / 11、20 / 11\20
TIME_LABEL_RE = re.compile(
    r'(\d{1,2}\s*(?:[./,，、\\]|-|－|—|。|．|／|月)+\s*\d{1,2})'
)

PRICE_MULTIPLIER = 4.7

# 同时支持半角/全角符号：> < = >= <= 以及 ＞ ＜ ＝ ≥ ≤
AMOUNT_TAG_RE = re.compile(
    r'开单金额\s*(?:>=|<=|>|<|=|＞|＜|＝|≥|≤)\s*(\d+(?:\.\d+)?)'
)

def calc_amount_from_tags(tag_str) -> float:
    """
    优先匹配：开单金额 + 比较符号 + 数字（半角/全角都支持）
    若没匹配到，则兜底：提取字符串里所有数字（防止写法不一致导致一直为0）
    """
    if tag_str is None:
        return 0.0

    s = str(tag_str).strip()
    if not s:
        return 0.0

    nums = []
    matches = AMOUNT_TAG_RE.findall(s)
    if matches:
        for m in matches:
            try:
                nums.append(float(m))
            except Exception:
                pass
    else:
        # 兜底：万一标签写法不是“开单金额>xxx”，也别一直算 0
        for m in re.findall(r'\d+(?:\.\d+)?', s):
            try:
                nums.append(float(m))
            except Exception:
                pass

    return sum(nums) * PRICE_MULTIPLIER

def get_black_delete_flags(tag_str):
    """
    返回访客标签中是否包含【黑粉】、【删/删除】。
    - “黑粉（澳）/黑粉(澳)”会命中黑粉
    - “删/删除”都会命中删除类关键词
    """
    if tag_str is None:
        return False, False

    s = str(tag_str)
    s = (
        s.replace("（", "(")
         .replace("）", ")")
         .replace("，", ",")
         .replace(" ", "")
         .strip()
    )

    has_black = "黑粉" in s
    has_delete = "删" in s  # 同时兼容“删”和“删除”
    return has_black, has_delete

def is_exclude_high_intent_by_tag(tag_str) -> bool:
    """
    正常高意向排除规则：
    - 只要标签包含【黑粉】或【删/删除】任意一个关键词，
      就不统计进正常高意向，而归入【高意向黑/删】。
    - 黑粉和删/删除同时存在时，也归入【高意向黑/删】。
    """
    has_black, has_delete = get_black_delete_flags(tag_str)
    return has_black or has_delete

def is_high_intent_black_delete_by_tag(tag_str) -> bool:
    """
    汇总表【高意向黑粉/删】统计规则：
    统计“本来满足高意向，且标签中包含黑粉或删/删除任意一个关键词”的记录。
    """
    has_black, has_delete = get_black_delete_flags(tag_str)
    return has_black or has_delete


def extract_time_labels(text: str):
    if not isinstance(text, str):
        return set()
    matches = TIME_LABEL_RE.findall(text)
    return set(matches)

def _normalize_md_label(label: str):
    """
    将各种分隔符的日期（11.3 / 11/03 / 11月3 / 11-20 / 12,20 / 12、20 / 12\20）统一转成 'MM-DD'
    """
    if not isinstance(label, str):
        return None
    # 注意：这里的正则也应该同步更新，或者复用 TIME_LABEL_RE 的核心部分
    # 为了保持一致性，我们直接复用 TIME_LABEL_RE 的模式，去掉捕获组的括号
    # pattern_core = r'\d{1,2}\s*[./,\-\\月、]\s*\d{1,2}'
    # m = re.search(pattern_core, label)
    
    # 或者，更简单地，直接搜索 TIME_LABEL_RE 匹配的内容
    m = TIME_LABEL_RE.search(label) # search 整个匹配
    if not m:
        return None
    # 从匹配的整个字符串中再次提取月份和日期数字
    # 使用一个通用的提取数字的正则
    numbers = re.findall(r'\d{1,2}', m.group(1)) # m.group(1) 是第一个捕获组的内容
    if len(numbers) < 2:
        return None
    try:
        # 假设第一个数字是月份，第二个是日期
        month = int(numbers[0]) 
        day = int(numbers[1])
        if not (1 <= month <= 12 and 1 <= day <= 31):
            return None
        return f"{month:02d}-{day:02d}"
    except (ValueError, IndexError):
        return None

# ========= whatsapp 覆盖逻辑（按联系人维度） =========
def apply_whatsapp_override(df):
    """
    规则：
    - 先按【联系人】聚合所有访客标签
    - 若聚合标签中含 whatsapp，则：
        · 含 全款/定金/分期/重点 -> 该联系人所有记录 综合标记 = 重点
        · 含 一般               -> 综合标记 = 1
        · 含 0互动              -> 综合标记 = 0
    - 同步覆盖“期望访客标签（澳）”（如果该列存在）

    新增约束：
    - 若该联系人属于全局 deal_three_labels_contacts（即：成交 + ≥3 时间标签）
      则跳过 whatsapp 覆盖逻辑（保持原有 0/1/重点 判定），防止被 whatsapp 强制改成重点或一般。
    """
    if visitor_tag_col not in df.columns or contact_col not in df.columns:
        return df

    # 每个联系人的所有标签合并到一段字符串
    tags_by_contact = df.groupby(contact_col)[visitor_tag_col] \
                        .apply(lambda s: "，".join(s.dropna().astype(str))) \
                        .astype(str).str.lower()

    contact_to_mark = {}
    for contact, tags_lower in tags_by_contact.items():
        # 3 时间标签 + 成交 的联系人，直接跳过 whatsapp 覆盖
        if str(contact) in deal_three_labels_contacts:
            continue

        if "whatsapp" not in tags_lower and "message" not in tags_lower:
            continue

        mark = None
        if re.search("全款|定金|分期|重点", tags_lower):
            mark = "重点"
        elif "一般" in tags_lower:
            mark = 1
        elif "0互动" in tags_lower:
            mark = 0

        if mark is not None:
            contact_to_mark[str(contact)] = mark

    if not contact_to_mark:
        return df

    df_contact_str = df[contact_col].astype(str)
    has_expect_col = "期望访客标签（澳）" in df.columns

    for c, mark in contact_to_mark.items():
        mask = (df_contact_str == c)
        df.loc[mask, "客户发言综合标记"] = mark

        if has_expect_col:
            if mark == "重点":
                df.loc[mask, "期望访客标签（澳）"] = "重点"
            elif mark == 1:
                df.loc[mask, "期望访客标签（澳）"] = "一般客户"
            elif mark == 0:
                df.loc[mask, "期望访客标签（澳）"] = "0互动"

    return df

# ============ 拖放文件选择窗口 ============
class FileDropWindow(Toplevel):
    """
    选择文件窗口：支持“按钮选择”，如系统安装了 tkinterdnd2 还支持“拖拽文件到窗口”。
    - file_count=1：选择 1 个文件
    - file_count=2：选择 2 个文件（按加入顺序作为 file1 / file2）
    """
    def __init__(self, parent, title, file_count=1, dnd_enabled=True):
        super().__init__(parent)
        self.parent = parent
        self.title(title)
        self.geometry("560x260")
        self.minsize(520, 240)

        self.file_paths = []
        self.file_count = int(file_count)
        self.selected = False
        self.dnd_enabled = bool(dnd_enabled)

        # 拖放区域
        self.drop_frame = Frame(self, bd=2, relief=tk.SUNKEN)
        self.drop_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=(20, 10))

        tip = (
            f"请拖放 {self.file_count} 个 Excel 文件到此处\\n"
            f"或点击下方按钮选择文件\\n\\n"
            f"支持：.xlsx / .xls"
        )
        if not self.dnd_enabled:
            tip += "\n\n（提示：未检测到 tkinterdnd2，仅可点击选择文件）"
        self.label = tk.Label(self.drop_frame, text=tip, justify="center")
        self.label.pack(expand=True)

        # 仅在支持拖放时注册 DND
        if self.dnd_enabled:
            try:
                self.drop_frame.drop_target_register(DND_FILES)  # type: ignore
                self.drop_frame.dnd_bind('<<Drop>>', self.on_drop)  # type: ignore
            except Exception:
                # 某些环境下 tkinterdnd2 存在但初始化失败，退回到“仅按钮选择”
                self.dnd_enabled = False
                self.label.config(text=tip + "\n\n（提示：拖放初始化失败，已退回按钮选择）")

        self.btn_select = tk.Button(self, text="选择文件", command=self.select_files)
        self.btn_select.pack(pady=(0, 8))

        self.file_listbox = tk.Listbox(self, height=min(6, self.file_count))
        self.file_listbox.pack(fill=tk.X, padx=20, pady=(0, 8))

        self.btn_confirm = tk.Button(self, text="确认", command=self.confirm)
        self.btn_confirm.pack(pady=(0, 12))

        # 使窗口居中（相对 parent）
        try:
            self.update_idletasks()
            pw = parent.winfo_width()
            ph = parent.winfo_height()
            px = parent.winfo_rootx()
            py = parent.winfo_rooty()
            w = self.winfo_width()
            h = self.winfo_height()
            x = px + (pw - w) // 2
            y = py + (ph - h) // 2
            self.geometry(f"{w}x{h}+{x}+{y}")
        except Exception:
            pass

    def _add_files(self, files):
        for file in files:
            file = str(file).strip().strip("{}")  # 兼容某些拖拽返回 {path}
            if not file:
                continue
            if not file.lower().endswith(('.xlsx', '.xls')):
                continue
            if len(self.file_paths) >= self.file_count:
                break
            if file not in self.file_paths:
                self.file_paths.append(file)
                self.file_listbox.insert(tk.END, os.path.basename(file))

        if len(self.file_paths) == self.file_count:
            self.confirm()

    def on_drop(self, event):
        try:
            files = self.tk.splitlist(event.data)
        except Exception:
            files = [event.data]
        self._add_files(files)

    def select_files(self):
        files = filedialog.askopenfilenames(
            title=f"选择 {self.file_count} 个 Excel 文件",
            filetypes=[["Excel", "*.xlsx;*.xls"], ["All files", "*.*"]]
        )
        if files:
            self._add_files(files)

    def confirm(self):
        if len(self.file_paths) == self.file_count:
            self.selected = True
            self.destroy()
        else:
            messagebox.showwarning("提示", f"请至少选择 {self.file_count} 个文件。")

# ============ GUI 调试工具 ============
def launch_debug_viewer(detail_df, deal_contacts_set, history_focus_contacts_set, root=None):
    """
    联系人明细调试窗口：
    - 选择接待成员
    - 选择类型（0互动 / 一般 / 重点 / 成交 / 历史重点提升 / 低意向客户 / 全部）
    - 查看对应联系人的明细，包括 客户发言数 + 客服发言数
    """
    if root is None:
        root = Tk()
        root.withdraw()

    win = Toplevel(root)
    win.title("联系人明细调试工具")
    win.geometry("900x600")

    frm_top = tk.Frame(win)
    frm_top.pack(side=tk.TOP, fill=tk.X, padx=10, pady=10)

    tk.Label(frm_top, text="接待成员：").grid(row=0, column=0, sticky="w", padx=5, pady=5)
    agent_list = sorted(detail_df[agent_col].dropna().astype(str).unique().tolist())
    cmb_agent = ttk.Combobox(frm_top, values=agent_list, state="readonly", width=18)
    cmb_agent.grid(row=0, column=1, sticky="w", padx=5, pady=5)

    tk.Label(frm_top, text="类型：").grid(row=0, column=2, sticky="w", padx=5, pady=5)
    type_options = ["全部", "0互动", "一般", "重点", "高意向客户", "高意向黑/删", "成交", "历史重点提升"]
    cmb_type = ttk.Combobox(frm_top, values=type_options, state="readonly", width=18)
    cmb_type.current(0)
    cmb_type.grid(row=0, column=3, sticky="w", padx=5, pady=5)

    btn_query = tk.Button(frm_top, text="查看联系人细节")
    btn_query.grid(row=0, column=4, sticky="w", padx=10, pady=5)

    frm_bottom = tk.Frame(win)
    frm_bottom.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=10, pady=5)

    txt_result = tk.Text(frm_bottom, wrap="word")
    txt_result.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    scroll = tk.Scrollbar(frm_bottom, command=txt_result.yview)
    scroll.pack(side=tk.RIGHT, fill=tk.Y)
    txt_result.config(yscrollcommand=scroll.set)

    def filter_by_type(df_sub, t_name):
        df = df_sub.copy()
        mark_col = df.get("客户发言综合标记")
        eff_col = df.get("客户发言数")
        if mark_col is not None:
            mark_col = mark_col.fillna("")
        if eff_col is not None:
            eff_col = eff_col.fillna(0)

        if t_name == "全部":
            return df

        # 0互动：排除 3 时间标签成交客户
        if t_name == "0互动":
            if mark_col is not None:
                mask = (mark_col == 0)
                if deal_three_labels_contacts:
                    mask = mask & ~df[contact_col].astype(str).isin(deal_three_labels_contacts)
                return df[mask]
            if eff_col is not None:
                mask = (eff_col <= 0)
                if deal_three_labels_contacts:
                    mask = mask & ~df[contact_col].astype(str).isin(deal_three_labels_contacts)
                return df[mask]
            return df.iloc[0:0]

        # 一般：排除 3 时间标签成交客户
        if t_name == "一般":
            if mark_col is not None:
                mask = (mark_col == 1)
                if deal_three_labels_contacts:
                    mask = mask & ~df[contact_col].astype(str).isin(deal_three_labels_contacts)
                return df[mask]
            return df.iloc[0:0]

        if t_name == "重点":
            if mark_col is not None:
                # 原本：重点但不是历史重点的记录
                mask = (mark_col == "重点") & ~df[contact_col].astype(str).isin(history_focus_contacts_set)
                # 新增：剔除 3 时间标签 + 成交 的联系人
                if deal_three_labels_contacts:
                    mask = mask & ~df[contact_col].astype(str).isin(deal_three_labels_contacts)
                return df[mask]
            return df.iloc[0:0]

        if t_name == "高意向客户":
            if (mark_col is not None) and (eff_col is not None):
                high_exclude = df[visitor_tag_col].apply(is_exclude_high_intent_by_tag)
                mask = (mark_col == 1) & (eff_col >= 6) & (~high_exclude)
                if deal_three_labels_contacts:
                    mask = mask & ~df[contact_col].astype(str).isin(deal_three_labels_contacts)
                return df[mask]
            return df.iloc[0:0]

        if t_name == "高意向黑/删":
            if (mark_col is not None) and (eff_col is not None):
                high_black_delete = df[visitor_tag_col].apply(is_high_intent_black_delete_by_tag)
                mask = (mark_col == 1) & (eff_col >= 6) & high_black_delete
                if deal_three_labels_contacts:
                    mask = mask & ~df[contact_col].astype(str).isin(deal_three_labels_contacts)
                return df[mask]
            return df.iloc[0:0]


        if t_name == "成交":
            mask = df[contact_col].astype(str).isin(deal_contacts_set)
            return df[mask]

        if t_name == "历史重点提升":
            mask = df[contact_col].astype(str).isin(history_focus_contacts_set)
            return df[mask]

        return df

    def on_query():
        agent = cmb_agent.get().strip()
        t_name = cmb_type.get().strip()

        if not agent:
            messagebox.showwarning("提示", "请先选择接待成员。")
            return

        sub = detail_df[detail_df[agent_col].astype(str) == agent].copy()
        if sub.empty:
            txt_result.delete("1.0", tk.END)
            txt_result.insert(tk.END, f"接待成员【{agent}】没有任何记录。\n")
            return

        sub = filter_by_type(sub, t_name)

        txt_result.delete("1.0", tk.END)

        if sub.empty:
            txt_result.insert(tk.END, f"接待成员【{agent}】在类型【{t_name}】下没有匹配的联系人。\n")
            return

        contacts_unique = sub[contact_col].astype(str).nunique()
        txt_result.insert(
            tk.END,
            f"=== 接待成员【{agent}】｜类型【{t_name}】｜"
            f"共 {contacts_unique} 位客户（{len(sub)} 条记录）===\n\n"
        )

        for _, row in sub.iterrows():
            contact = str(row.get(contact_col, ""))
            tags = str(row.get(visitor_tag_col, ""))
            mark = str(row.get("客户发言综合标记", ""))
            eff = row.get("客户发言数", "")
            eff_agent = row.get("客服发言数", "")

            line = (
                f"联系人：{contact}\n"
                f"  客户发言综合标记：{mark}    客户发言数：{eff}    客服发言数：{eff_agent}\n"
                f"  访客标签：{tags}\n"
                "------------------------------------------------------------\n"
            )
            txt_result.insert(tk.END, line)

    btn_query.config(command=on_query)

    return win

# ======================= 主流程 =======================
def _create_root():
    """创建主窗口：若安装了 tkinterdnd2 则启用拖放，否则回退普通 Tk。"""
    if _DND_AVAILABLE and TkinterDnD is not None:
        try:
            r = TkinterDnD.Tk()  # type: ignore
            return r, True
        except Exception:
            pass
    r = Tk()
    return r, False


def _make_unique_path(p: str) -> str:
    """若同名文件已存在，则自动追加 _01/_02... 防止覆盖。"""
    base, ext = os.path.splitext(p)
    if not os.path.exists(p):
        return p
    idx = 1
    while True:
        cand = f"{base}_{idx:02d}{ext}"
        if not os.path.exists(cand):
            return cand
        idx += 1


def load_and_merge_excel_files(file1: str, file2: str) -> pd.DataFrame:
    """读取两个 Excel 并纵向合并（与原先手动导出流程一致）。"""
    df1 = read_any_excel(file1)
    df2 = read_any_excel(file2)
    df1.columns = df1.columns.str.strip()
    df2.columns = df2.columns.str.strip()
    final_columns = list(df1.columns) + [c for c in df2.columns if c not in df1.columns]
    merged_df = pd.concat([df1, df2], ignore_index=True, join="outer", sort=False)
    merged_df = merged_df.reindex(columns=final_columns)
    merged_df.columns = merged_df.columns.str.strip()
    return merged_df


def run_analysis(merged_df: pd.DataFrame, save_dir: str, root) -> None:
    """对合并后的数据执行分析，输出标签详情明细与汇总数据。"""
    try:
        if root is not None:
            root.deiconify()
            root.lift()
    except Exception:
        pass

    os.makedirs(save_dir, exist_ok=True)

    global contact_col, session_col, message_col, visitor_tag_col, agent_col
    global deal_three_labels_contacts

    today_str = datetime.now().strftime("%Y%m%d")
    output_file1 = _make_unique_path(os.path.join(save_dir, f"标签详情明细_{today_str}.xlsx"))
    output_file2 = _make_unique_path(os.path.join(save_dir, f"汇总数据_{today_str}.xlsx"))

    merged_df = merged_df.copy()
    merged_df.columns = merged_df.columns.str.strip()

    # 明细表（合并前的详细数据）
    detail_df = merged_df.copy()

    # 6) 智能定位关键列
    contact_col = find_column(
        merged_df.columns,
        ["联系人", "客户", "用户", "contact", "user", "客户姓名", "联系人姓名"],
        contains=True
    )
    session_col  = find_column(
        merged_df.columns,
        ["会话id", "会话ID", "sessionid", "session_id", "会话编号"],
        contains=True
    )
    message_col  = find_column(
        merged_df.columns,
        ["会话消息内容", "会话内容", "聊天内容", "对话内容", "消息内容", "消息", "文本", "text", "message"],
        contains=True
    )
    visitor_tag_col = find_column(
        merged_df.columns,
        ["访客标签", "客户标签", "标签", "visitor tag", "customer tag", "tag"],
        contains=True
    )
    agent_col    = find_column(
        merged_df.columns,
        ["接待成员", "接待客服", "接待人员", "接待员",  "客服", "客服姓名", "客服名称", "坐席", "坐席名", "agent", "staff"],
        contains=True
    )

    missing = []
    if not contact_col: missing.append("联系人(如：联系人/客户/用户等)")
    if not session_col: missing.append("会话id(如：会话ID/sessionid 等)")
    if not message_col: missing.append("会话消息内容(如：会话消息内容/会话内容/消息内容/聊天内容 等)")
    if not visitor_tag_col: missing.append("访客标签(如：访客标签/客户标签/标签 等)")
    if not agent_col: missing.append("接待成员/接待客服(如：接待成员/客服/客服姓名 等)")
    if missing:
        raise ValueError(
            "未能自动匹配到以下必需列，请检查列名或扩充匹配规则：\n - " +
            "\n - ".join(missing)
        )
    session_tag_col = "会话标签"
    if session_tag_col not in merged_df.columns:
        raise ValueError("未找到列【会话标签】。请确认Excel里列名严格为：会话标签")
    
    # ===== 根据【接待成员】列动态重建“客服发言”匹配规则 =====
    # 让 [时间] 王栋栋 : “内容” 这类发言，也被当成客服发言
    if agent_col:
        raw_agents = merged_df[agent_col].dropna().astype(str)
        agent_name_set = set()

        for val in raw_agents:
            for part in re.split(r'[，,、/|;]', val):
                name = part.strip()
                if name:
                    agent_name_set.add(name)

        if agent_name_set:
            update_agent_re(agent_name_set)
    # 7) 结果页聚合
    def join_unique(series, sep):
        vals = [str(x).strip() for x in series.dropna().astype(str) if str(x).strip() != ""]
        if not vals:
            return None
        seen, out = set(), []
        for v in vals:
            if v not in seen:
                seen.add(v)
                out.append(v)
        return sep.join(out)

    agg_map = {col: first_non_null for col in merged_df.columns}
    agg_map[message_col] = lambda s: join_unique(s, "\n")
    agg_map[visitor_tag_col] = lambda s: join_unique(s, "，")
    agg_map[agent_col] = lambda s: join_unique(s, "、")

    merged_df = merged_df.groupby([contact_col, session_col], as_index=False).agg(agg_map)

    # 8) 结果页标记（客户 + 客服发言数）
    merged_df["客户发言综合标记"] = merged_df[message_col].apply(combined_marker)
    merged_df["客户发言数"] = merged_df[message_col].apply(count_valid_customer_segments)
    merged_df["客服发言数"] = merged_df[message_col].apply(count_valid_agent_segments)

    # 宽泛类别映射：去掉（澳）
    mapping = {0: "0互动", 1: "一般客户", "重点": "重点"}
    merged_df["期望访客标签（澳）"] = merged_df["客户发言综合标记"].map(mapping).fillna("")

    # ========== 成交客户（按联系人）【先算是否成交列，后面所有逻辑都用这列】 ==========
    deal_pattern = "全款|定金|分期"
    vtag_lower = merged_df[visitor_tag_col].fillna("").astype(str).str.lower()
    merged_df["是否成交"] = vtag_lower.str.contains(deal_pattern, na=False)
    contact_deal_status = merged_df.groupby(contact_col)["是否成交"].any()
    deal_contacts = set(contact_deal_status[contact_deal_status].index.astype(str))

    # ====== 预先计算：3 个及以上时间标签 + 成交 的联系人集合，用于后续 whatsapp 覆盖与统计 ======
    tag_by_contact_pre = (
        merged_df.groupby(contact_col)[visitor_tag_col]
                 .apply(lambda s: "，".join(s.dropna().astype(str)))
    )
    deal_three_labels_contacts = set()
    for contact, tags_str in tag_by_contact_pre.items():
        labels_raw = list(extract_time_labels(tags_str))
        norm_labels = []
        for lb in labels_raw:
            md = _normalize_md_label(lb)
            if md:
                norm_labels.append(md)
        if not norm_labels:
            continue

        # 使用上面算好的 contact_deal_status 判断是否成交
        is_deal = bool(contact_deal_status.get(contact, False))
        if is_deal and len(norm_labels) >= 3:
            # 成交 + 3 个及以上时间标签 -> 只算成交
            deal_three_labels_contacts.add(str(contact))

    # ===== whatsapp 按联系人覆盖：0/1/重点 按访客标签决定（但跳过 deal_three_labels_contacts）=====
    merged_df = apply_whatsapp_override(merged_df)

    # 高意向客户规则：
    # - 原规则：综合标记=1 且 客户发言数>=6
    # - 如果访客标签包含“黑粉”或“删/删除”任意一个，则期望标签改为【高意向黑/删】
    # - 不含黑粉/删/删除时，才归入正常【高意向客户】
    base_high_intent_series = (
        (merged_df["客户发言综合标记"] == 1) &
        (merged_df["客户发言数"] >= 6)
    )
    high_intent_black_delete_series = merged_df[visitor_tag_col].apply(is_high_intent_black_delete_by_tag)
    high_intent_mask_series = base_high_intent_series & (~high_intent_black_delete_series)
    high_intent_black_delete_mask_series = base_high_intent_series & high_intent_black_delete_series

    merged_df.loc[high_intent_mask_series, "期望访客标签（澳）"] = "高意向客户"
    merged_df.loc[high_intent_black_delete_mask_series, "期望访客标签（澳）"] = "高意向黑/删"

    # ===== 用【客户发言综合标记】 vs 访客标签关键词 判定是否标红 =====
    focus_aliases = ["重点", "全款", "定金", "分期"]

    def tag_to_mark(tag_str):
        """
        根据访客标签内容推断“标签对应的类型”：
        - 含 0互动       -> 0
        - 含 重点/全款/定金/分期 -> "重点"
        - 含 一般        -> 1
        - 其它情况       -> None（不判定，不参与标红）
        """
        s = (tag_str or "").lower().replace(" ", "")
        if not s:               # 标签为空，也视作 0互动类
            return 0
        if "0互动" in s:
            return 0
        if any(k in s for k in focus_aliases):
            return "重点"
        if "一般" in s:
            return 1
        return None   # 没有明显关键词的标签，不用于判红

    expected_mark_series = merged_df[visitor_tag_col].fillna("").astype(str).apply(tag_to_mark)
    actual_mark_series   = merged_df["客户发言综合标记"]

    match_mask = []
    for actual, expected in zip(actual_mark_series.tolist(), expected_mark_series.tolist()):
        if expected is None:
            match_mask.append(True)   # 不判错
        else:
            match_mask.append(actual == expected)

    # ========== 历史重点提升（全局，按联系人 + 聊天内容一起判断） ==========
    tag_by_contact = (
        merged_df.groupby(contact_col)[visitor_tag_col]
                 .apply(lambda s: "，".join(s.dropna().astype(str)))
    )
    msgs_by_contact = (
        merged_df.groupby(contact_col)[message_col]
                 .apply(lambda s: list(s.dropna().astype(str)))
    )

    today_md = datetime.now().strftime("%m-%d")
    history_focus_contacts_global = set()

    for contact, tags_str in tag_by_contact.items():
        labels_raw = list(extract_time_labels(tags_str))
        if len(labels_raw) < 2:
            continue

        norm_labels = []
        for lb in labels_raw:
            md = _normalize_md_label(lb)
            if md:
                norm_labels.append(md)
        if not norm_labels:
            continue

        # 是否成交，直接看前面算好的 set
        is_deal_contact = str(contact) in deal_contacts
        has_focus_tag = "重点" in tags_str

        msgs = msgs_by_contact.get(contact, [])
        has_focus_keyword = any(agent_focus_hit(m) for m in msgs)

        if is_deal_contact:
            # 3 个及以上时间标签 + 成交：只算成交，不算历史重点
            if len(norm_labels) >= 3:
                # 此类联系人已经在 deal_three_labels_contacts 中标记过
                continue

            # 只有 2 个时间标签 + 成交：
            #   需要：最晚日期 == 今天，且聊天里有重点关键词，才算历史重点提升
            if len(norm_labels) == 2:
                latest_md = max(norm_labels)
                if (latest_md == today_md) and has_focus_keyword:
                    history_focus_contacts_global.add(str(contact))
                continue
        else:
            # 不含成交标签：只要有“重点”且时间标签>=2
            if has_focus_tag and len(norm_labels) >= 2:
                history_focus_contacts_global.add(str(contact))
                continue

    # ===== 成交客户不参与标红：使用上面已经算好的 merged_df["是否成交"] =====
    deal_series = merged_df["是否成交"].fillna(False)

    mismatch_mask = []
    for ok, is_deal in zip(match_mask, deal_series):
        if is_deal:
            mismatch_mask.append(False)   # 成交客户绝不标红
        else:
            mismatch_mask.append(not ok)

    # 11) 写“结果”表 - 自动保存
    with pd.ExcelWriter(output_file1, engine="openpyxl") as writer:
        merged_df.to_excel(writer, index=False, sheet_name="结果")
        ws = writer.sheets["结果"]

        red_fill = PatternFill(fill_type="solid", start_color="FFFFC7CE", end_color="FFFFC7CE")
        yellow_fill = PatternFill(fill_type="solid", start_color="FFFFEB9C", end_color="FFFFEB9C")
        # 新增绿色填充样式（浅绿底色）
        green_fill = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")
        # 1. 定义黑色填充样式（放在其他填充样式定义处）
        black_fill = PatternFill(fill_type="solid", start_color="000000", end_color="000000")  # 纯黑色

        n_rows, n_cols = merged_df.shape

        # 按 mismatch_mask 标红整行
        for row_idx, mism in enumerate(mismatch_mask, start=2):
            if mism:
                for col_idx in range(1, n_cols + 1):
                    ws[f"{get_column_letter(col_idx)}{row_idx}"].fill = red_fill

        # 高意向客户联系人标黄（联系人这一格）
        contact_col_idx_excel = merged_df.columns.get_loc(contact_col) + 1
        for row_idx, is_high in enumerate(high_intent_mask_series.tolist(), start=2):
            if is_high:
                ws[f"{get_column_letter(contact_col_idx_excel)}{row_idx}"].fill = yellow_fill



        # 新增：访客标签无"一般""重点""0互动"时标绿（放在标红之后，确保绿色不被覆盖）
        visitor_tag_col_idx_excel = merged_df.columns.get_loc(visitor_tag_col) + 1
        for row_idx in range(2, n_rows + 2):  # 行号从2开始（表头为1）
            # 获取当前行的访客标签值
            tag_value = merged_df.iloc[row_idx - 2][visitor_tag_col]  # 转换为DataFrame的0-based索引
            tag_str = str(tag_value).lower()
            # 检查是否不包含三个关键词
            if "一般" not in tag_str and "重点" not in tag_str and "0互动" not in tag_str:
                # 标绿访客标签单元格
                ws[f"{get_column_letter(visitor_tag_col_idx_excel)}{row_idx}"].fill = green_fill

        if message_col in merged_df.columns:
            # 获取消息列的Excel列索引（从1开始）
            message_col_idx = merged_df.columns.get_loc(message_col) + 1  # 转换为Excel的1-based索引
            # 遍历所有数据行（Excel行号从2开始，对应merged_df的0-based索引）
            for row_idx in range(len(merged_df)):
                excel_row = row_idx + 2  # 数据行从第2行开始
                # 获取当前行的消息内容
                message_value = merged_df.iloc[row_idx][message_col]
                # 判定为空的条件：NaN、空字符串、仅空格
                is_empty = (pd.isna(message_value) or 
                        (isinstance(message_value, str) and message_value.strip() == ""))
                
                if is_empty:
                    # 遍历所有列，设置黑色填充
                    for col_idx in range(1, n_cols + 1):  # Excel列从1开始
                        cell = ws[f"{get_column_letter(col_idx)}{excel_row}"]
                        cell.fill = black_fill
                        # 强制设置字体颜色为白色（避免黑色背景看不到文字）
                        cell.font = cell.font.copy(color="FFFFFF")

    messagebox.showinfo("完成", f"合并完成，已保存结果表：\n{output_file1}")

    # 12) 生成【按接待成员统计】- 自动保存
    def load_department_mapping_from_excel(file_path):
        try:
            if os.path.exists(file_path):
                df_map = pd.read_excel(file_path)
                if "姓名" in df_map.columns and "部门" in df_map.columns:
                    return dict(zip(df_map["姓名"], df_map["部门"]))
        except Exception:
            pass
        return {}

    map_path = os.path.join(app_dir(), "人员部门表.xlsx")
    agent_to_department = load_department_mapping_from_excel(map_path)

    # 明细表上计算标记（坐席统计用）
    detail_df["客户发言综合标记"] = detail_df[message_col].apply(combined_marker)
    detail_df["客户发言数"] = detail_df[message_col].apply(count_valid_customer_segments)
    detail_df["客服发言数"] = detail_df[message_col].apply(count_valid_agent_segments)

    # 明细表也按联系人维度应用 whatsapp 覆盖，保证与结果表一致
    detail_df = apply_whatsapp_override(detail_df)

    # 关键修复：汇总统计直接使用【结果表】里已经最终判定好的【期望访客标签（澳）】。
    # 之前这里重新用明细表的访客标签计算，高意向已经在结果表改成【高意向黑/删】时，
    # 汇总表可能仍然统计不到【高意向黑粉/删】。
    expected_label_map_df = merged_df[[contact_col, session_col, "期望访客标签（澳）"]].copy()
    expected_label_map_df["_expected_key"] = (
        expected_label_map_df[contact_col].astype(str) + "||" + expected_label_map_df[session_col].astype(str)
    )
    expected_label_map = dict(
        zip(expected_label_map_df["_expected_key"], expected_label_map_df["期望访客标签（澳）"])
    )

    detail_df["_expected_key"] = detail_df[contact_col].astype(str) + "||" + detail_df[session_col].astype(str)
    detail_df["期望访客标签（澳）"] = detail_df["_expected_key"].map(expected_label_map).fillna("")
    detail_df.drop(columns=["_expected_key"], inplace=True)

    df_agents = detail_df.copy()
    df_agents[agent_col] = df_agents[agent_col].fillna("").astype(str).str.strip()
    df_agents = df_agents[df_agents[agent_col] != ""]
    df_agents[contact_col] = df_agents[contact_col].astype(str)

    df_agents[agent_col] = df_agents[agent_col].str.strip()

    df_agents[agent_col] = df_agents[agent_col].str.replace(r'[，,;/|]', '、', regex=True)

    exploded_rows = []
    for _, row in df_agents.iterrows():
        agents = [x.strip() for x in str(row[agent_col]).split('、') if x.strip()]
        for agent_name in agents:
            new_row = row.copy()
            new_row[agent_col] = agent_name
            exploded_rows.append(new_row)

    df_agents = pd.DataFrame(exploded_rows)
    df_agents = df_agents[df_agents[agent_col] != ""]
    # ====== 用 merged_df 先算出“每个接待员”的开单金额汇总 ======
    df_amount = merged_df[[agent_col, "会话标签"]].copy()
    df_amount["开单金额"] = df_amount["会话标签"].fillna("").astype(str).apply(calc_amount_from_tags)

    # merged_df 里的接待成员可能是 "张三、李四" 这种，拆开
    df_amount[agent_col] = df_amount[agent_col].fillna("").astype(str).str.replace(r'[，,;/|]', '、', regex=True)
    df_amount[agent_col] = df_amount[agent_col].apply(lambda x: [i.strip() for i in x.split("、") if i.strip()])
    
    df_amount = df_amount.explode(agent_col)

    amount_sum_by_agent = df_amount.groupby(agent_col)["开单金额"].sum().to_dict()

    # ====== 新增：每个接待员的 黑粉 / 删 客户数（按联系人去重）======
    df_black = merged_df[[agent_col, contact_col, visitor_tag_col]].copy()

    # 拆分接待成员
    df_black[agent_col] = (
        df_black[agent_col]
        .fillna("")
        .astype(str)
        .str.replace(r"[，,;/|]", "、", regex=True)
        .str.split("、")
    )
    df_black = df_black.explode(agent_col)
    df_black = df_black[df_black[agent_col].str.strip() != ""]

    # 统一标签文本
    tags_norm = (
        df_black[visitor_tag_col]
        .fillna("")
        .astype(str)
        .str.replace("（", "(", regex=False)
        .str.replace("）", ")", regex=False)
        .str.replace("，", ",", regex=False)
        .str.replace(" ", "", regex=False)
        .str.strip()
    )

    # 命中 黑粉（澳） 或 删
    black_mask = tags_norm.str.contains("黑粉", regex=False) | tags_norm.str.contains("删", regex=False)

    df_black_hit = df_black.loc[black_mask]

    # 按 接待员 + 联系人 去重计数
    black_count_by_agent = (
        df_black_hit
        .drop_duplicates(subset=[agent_col, contact_col])
        .groupby(agent_col)[contact_col]
        .nunique()
        .to_dict()
    )


    grouped = []

    for agent_name, sub in df_agents.groupby(agent_col):
        sub = sub.copy()
        sub[contact_col] = sub[contact_col].astype(str)

        sub_contacts = sub[contact_col].dropna().astype(str)
        contacts_all = set(sub_contacts.unique())
        
       

        contacts_history_focus = contacts_all & history_focus_contacts_global
        contacts_deal = contacts_all & deal_contacts

        is_0_sub          = (sub["客户发言综合标记"] == 0)
        is_1_sub          = (sub["客户发言综合标记"] == 1)
        is_focus_sub_raw  = (sub["客户发言综合标记"] == "重点")
        is_history_focus_record = sub[contact_col].astype(str).isin(contacts_history_focus)

        # 3 时间标签 + 成交 的联系人在 0/1/重点 统计中全部剔除
        if deal_three_labels_contacts:
            is_three_deal_record = sub[contact_col].astype(str).isin(deal_three_labels_contacts)
        else:
            is_three_deal_record = pd.Series(False, index=sub.index)

        # 0 / 1 / 重点，全部排除三时间成交客户
        is_0_sub_use     = is_0_sub & ~is_three_deal_record
        is_1_sub_use     = is_1_sub & ~is_three_deal_record
        is_focus_sub     = is_focus_sub_raw & ~is_history_focus_record & ~is_three_deal_record

        # 高意向统计直接读取【期望访客标签（澳）】最终值，确保和结果表一致。
        # 结果表已经把高意向分成【高意向客户】和【高意向黑/删】，这里不再重复用访客标签二次判断。
        expected_label_sub = sub["期望访客标签（澳）"].fillna("").astype(str).str.strip()
        is_high_sub = (expected_label_sub == "高意向客户") & ~is_three_deal_record
        is_high_black_delete_sub = (expected_label_sub == "高意向黑/删") & ~is_three_deal_record

        count_0      = int(sub.loc[is_0_sub_use].shape[0])
        count_1      = int(sub.loc[is_1_sub_use].shape[0])
        count_focus  = int(sub.loc[is_focus_sub].shape[0])
        count_high = int(sub.loc[is_high_sub].shape[0])
        count_high_black_delete = int(sub.loc[is_high_black_delete_sub].shape[0])

        total        = count_0 + count_1 + count_focus

        # ====== 新增：开单金额（按接待员汇总，不去重）======
        amount_total = float(amount_sum_by_agent.get(agent_name, 0.0))
        black_total = int(black_count_by_agent.get(agent_name, 0))



        grouped.append({
            "姓名": agent_name,
            "总数": total,
            "0互动": count_0,
            "一般": count_1,
            "重点": count_focus,
            "成交": len(contacts_deal),
            "历史重点提升": len(contacts_history_focus),
            "高意向客户": count_high,
            "高意向黑粉/删": count_high_black_delete,
            "开单金额": amount_total,
            "黑粉/删": black_total,

        })

    grouped_sorted = sorted(grouped, key=lambda d: (-d["总数"], d["姓名"]))

    # 保存统计结果
    wb_new = Workbook()
    ws_new = wb_new.active
    ws_new.title = "按接待成员统计"

    headers = [
        "填写日期时间",
        "部门单选",
        "成员单选",
        "进线人数",
        "0互动用户",
        "一般用户",
        "重点用户",
        "成交人数",
        "历史重点提升",
        "高意向客户",
        "高意向黑粉/删",
        "开单金额",
        "黑粉/删"
    ]

    ws_new.append(headers)


    current_date = datetime.now().strftime("%Y-%m-%d")
    for row in grouped_sorted:
        name = row["姓名"]
        dept = agent_to_department.get(name, "未知部门")
        ws_new.append([
            current_date,                    # 填写日期时间
            dept,                            # 部门单选
            name,                            # 成员单选
            int(row["总数"]),                # 进线人数
            int(row["0互动"]),               # 0互动用户
            int(row["一般"]),                # 一般用户
            int(row["重点"]),                # 重点用户
            int(row["成交"]),                # 成交人数
            int(row["历史重点提升"]),        # 历史重点提升
            int(row["高意向客户"]),          # 高意向客户
            int(row["高意向黑粉/删"]),    # 高意向黑粉/删
            round(float(row["开单金额"]), 2),# 开单金额
            int(row["黑粉/删"]),             # 黑粉/删
        ])

        
    wb_new.save(output_file2)
    messagebox.showinfo("完成", f"已生成新的【按接待成员统计】表：\n{output_file2}")

    # 是否打开调试窗口
    if messagebox.askyesno("调试工具", "是否打开【联系人细节调试】窗口（按接待成员 + 类型查看客户明细）？"):
        launch_debug_viewer(detail_df, deal_contacts, history_focus_contacts_global, root)
        root.mainloop()


def _show_error_dialog(title: str, message: str) -> None:
    """显示错误弹窗（不依赖已隐藏的 root）。"""
    r = Tk()
    r.withdraw()
    try:
        r.attributes("-topmost", True)
    except Exception:
        pass
    messagebox.showerror(title, message, parent=r)
    r.destroy()


def _write_error_log(message: str) -> str:
    log_dir = os.path.join(app_dir(), "输出")
    os.makedirs(log_dir, exist_ok=True)
    log_path = os.path.join(log_dir, "error.log")
    with open(log_path, "a", encoding="utf-8") as f:
        f.write(f"\n{'=' * 60}\n")
        f.write(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(message)
        if not message.endswith("\n"):
            f.write("\n")
    return log_path


def run_with_selection(selection) -> None:
    """根据筛选结果拉数并生成报表（重模块在此之后才加载）。"""
    import threading
    import traceback

    from filter_window import ProgressDialog
    from fetch_progress import FetchCancelledError, FetchProgressReporter
    from fetch_salesmartly import build_merged_dataframe
    from salesmartly_client import SaleSmartlyClient, load_config

    root = None

    try:
        root, _dnd_enabled = _create_root()
        root.withdraw()

        save_dir = os.path.join(app_dir(), "输出")

        root.deiconify()
        root.update_idletasks()

        progress = ProgressDialog(root, "正在从 SaleSmartly 拉取数据")
        pending_progress: list = []

        def _on_progress_ui(snap):
            pending_progress.append(snap)

        fetch_progress = FetchProgressReporter(ui_hook=_on_progress_ui)
        fetch_progress.bind_cancel_check(lambda: progress.cancelled)
        progress.set_on_cancel(fetch_progress.request_cancel)

        def _flush_progress_ui():
            while pending_progress:
                snap = pending_progress.pop(0)
                if progress.winfo_exists():
                    progress.apply_snapshot(
                        snap.stage, snap.current, snap.total, snap.detail
                    )

        result: dict = {}
        fetch_error: dict = {}

        def _fetch_worker():
            try:
                client = SaleSmartlyClient(load_config())
                result["df"] = build_merged_dataframe(
                    client,
                    selection.agent_ids,
                    selection.member_id_to_name,
                    selection.tag_names,
                    progress=fetch_progress,
                )
            except Exception as exc:
                fetch_error["exc"] = exc

        worker = threading.Thread(target=_fetch_worker, daemon=True)
        worker.start()

        while worker.is_alive():
            _flush_progress_ui()
            root.update()
            if progress.cancelled:
                fetch_progress.request_cancel()
            worker.join(timeout=0.05)

        _flush_progress_ui()

        try:
            progress.destroy()
        except Exception:
            pass

        if "exc" in fetch_error:
            raise fetch_error["exc"]
        if progress.cancelled and "df" not in result:
            messagebox.showinfo("已取消", "拉取已终止，未生成报表。", parent=root)
            return

        run_analysis(result["df"], save_dir, root)

    except FetchCancelledError:
        try:
            if root is not None:
                messagebox.showinfo("已取消", "拉取已终止，未生成报表。", parent=root)
            else:
                _show_error_dialog("已取消", "拉取已终止，未生成报表。")
        except Exception:
            pass
    except Exception as e:
        err_text = f"{type(e).__name__}: {e}"
        full_trace = traceback.format_exc()
        log_path = _write_error_log(full_trace)
        display = f"{err_text}\n\n详细日志已写入：\n{log_path}"
        try:
            if root is not None:
                try:
                    root.deiconify()
                    root.lift()
                except Exception:
                    pass
                messagebox.showerror("运行出错", display, parent=root)
            else:
                _show_error_dialog("运行出错", display)
        except Exception:
            print(full_trace)
    finally:
        if root is not None:
            try:
                root.destroy()
            except Exception:
                pass


def main():
    from filter_window import FilterWindow

    os.chdir(app_dir())
    if app_dir() not in sys.path:
        sys.path.insert(0, app_dir())

    filter_win = FilterWindow()
    filter_win.mainloop()

    if filter_win.selected is None:
        return

    run_with_selection(filter_win.selected)


if __name__ == "__main__":
    try:
        main()
    except Exception:
        import traceback
        try:
            _r = Tk()
            _r.withdraw()
            messagebox.showerror("启动失败", traceback.format_exc(), parent=_r)
            _r.destroy()
        except Exception:
            traceback.print_exc()