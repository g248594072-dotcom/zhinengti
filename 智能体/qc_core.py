# -*- coding: utf-8 -*-
"""
聊天质检 · 核心逻辑（桌面版 / 网页版共用）
"""

import os
import sys
import json
import re
import time
import logging
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed

import pandas as pd
import requests

logger = logging.getLogger(__name__)

# ============================================================
# Prompt 加载
# ============================================================
_PROMPT_CACHE = {}


def get_app_dir():
    """程序所在目录（兼容 .py、PyInstaller exe、中文路径）。"""
    if getattr(sys, "frozen", False):
        return os.path.dirname(os.path.abspath(sys.executable))
    # qc_core.py 与 聊天质检工具.py 同目录
    return os.path.dirname(os.path.abspath(__file__))


def get_config_dir():
    """配置目录：Docker 可通过环境变量 QC_CONFIG_DIR 指向挂载卷（如 /app/config）。"""
    override = (os.environ.get("QC_CONFIG_DIR") or "").strip()
    if override:
        return override
    return get_app_dir()


def _prompt_search_paths(filename):
    app_dir = get_app_dir()
    paths = [os.path.join(app_dir, "prompts", filename)]
    meipass = getattr(sys, "_MEIPASS", None)
    if meipass:
        paths.append(os.path.join(meipass, "prompts", filename))
    return paths


def load_prompt(filename):
    if filename in _PROMPT_CACHE:
        return _PROMPT_CACHE[filename]
    for path in _prompt_search_paths(filename):
        if os.path.isfile(path):
            with open(path, "r", encoding="utf-8") as f:
                content = f.read().strip()
            _PROMPT_CACHE[filename] = content
            return content
    tried = "\n".join(f"  · {p}" for p in _prompt_search_paths(filename))
    raise FileNotFoundError(
        f"无法加载 Prompt「{filename}」。\n已尝试路径：\n{tried}\n\n"
        f"请确认「prompts」文件夹与程序在同一目录。"
    )


def enrich_system_prompt(base: str, tier: str = "full") -> str:
    """合并已审核的成交复盘规则补充（prompts/deal_learned_supplement.md）。"""
    if tier in ("deep_review", "progress_compare"):
        return base
    try:
        sup = load_prompt("deal_learned_supplement.md")
    except FileNotFoundError:
        return base
    sup = re.sub(r"<!--[\s\S]*?-->", "", sup or "").strip()
    if not sup:
        return base
    return base + "\n\n# 成交复盘沉淀规则（已审核合并）\n" + sup


# ============================================================
# 分级与报告列
# ============================================================
TIER_SKIP_MAX = 3
TIER_LIGHT_MAX = 6
LITE_MAX_CHARS = 6000
UTTERANCE_GAP_MINUTES = 5
LOW_INTENT_REPLY_SLOW_MINUTES = 10
# 进度对比：销售触达按 10 分钟内连续多条合并为 1 次
PROGRESS_COMPARE_TOUCH_GAP_MINUTES = 10
# 低意向超时标红：仅当客户首次发言落在以下时段内才统计（左闭右开，按当日 clock time）
LOW_INTENT_REPLY_WINDOWS = (
    ((9, 30), (12, 0)),    # 09:30 ~ 12:00
    ((13, 30), (17, 0)),   # 13:30 ~ 17:00
    ((18, 30), (20, 0)),   # 18:30 ~ 20:00
)
LOW_INTENT_SPEECH_PAIRS = [
    ("客户第一次发言", "客服第一次发言"),
    ("客户第二次发言", "客服第二次发言"),
    ("客户第三次发言", "客服第三次发言"),
]
LOW_INTENT_SPEECH_COLUMNS = [col for pair in LOW_INTENT_SPEECH_PAIRS for col in pair]

TIME_SCOPE_ALL = "all"
TIME_SCOPE_TODAY = "today"
TIME_SCOPE_CUSTOM = "custom"
BUSINESS_DAY_CUTOFF_HOUR = 20

TIER_LABEL = {"skip": "跳过", "light": "轻量", "full": "重点"}

TIME_SCOPE_LABELS = {
    TIME_SCOPE_ALL: "全部",
    TIME_SCOPE_TODAY: "当天",
    TIME_SCOPE_CUSTOM: "自定义",
}

REPORT_COLUMNS = [
    "会话ID", "联系人", "接待成员", "渠道", "客户发言数", "客服发言数", "质检档位",
    "结果标签", "客户阶段", "客户主要顾虑",
    "是否合格", "风险等级", "是否需要人工复核",
    "红线", "对症率", "疑问及时率", "承接评价",
    "顾虑_效果", "顾虑_信任", "顾虑_价格", "顾虑_拖延", "顾虑_COD",
    "客户心理轨迹", "沉默与激活", "逼单跟单",
    "主要问题", "改进建议", "下一步跟进动作",
    "成交参考案例数", "成交参考摘要",
    "_错误", "原始对话",
]

LOW_INTENT_ANALYSIS_COLUMNS = [
    "快速筛选",
    "筛选说明",
    *LOW_INTENT_SPEECH_COLUMNS,
]

LOW_INTENT_REPORT_COLUMNS = [
    "会话ID", "联系人", "接待成员", "渠道",
    "客户发言数", "客服发言数",
    "快速筛选", "筛选说明",
    *LOW_INTENT_SPEECH_COLUMNS,
    "结果标签", "原始对话",
]

# 深度复盘：客户发言数 > 10 才纳入
DEEP_REVIEW_MIN_CUSTOMER_SPEECHES = 10
DEEP_REVIEW_PROMPT_FILE = "deep_review_qc.md"
REPORT_SHEET_DEEP_REVIEW = "深度复盘"

DEEP_REVIEW_REPORT_COLUMNS = [
    "会话ID", "联系人", "接待成员", "渠道", "客户发言数", "客服发言数",
    "结果标签", "客户当前阶段",
    "卡点诊断", "为什么不付钱", "卡住原因",
    "客户主要顾虑", "顾虑详解",
    "客户心理画像", "心理轨迹与转折", "未化解障碍",
    "客服已做与缺失", "推荐下一步动作", "跟进话术建议",
    "成交可能性", "紧迫度", "风险提醒",
    "_错误", "原始对话",
]

# 进度对比：同一天多份质检报告之间，按时间逐段比对同一客户的跟进进展
PROGRESS_COMPARE_PROMPT_FILE = "progress_compare.md"
REPORT_SHEET_PROGRESS_COMPARE = "进度对比"

PROGRESS_COMPARE_REPORT_COLUMNS = [
    "联系人", "会话ID", "接待成员", "渠道",
    "对比段", "较早报告", "较晚报告",
    "较早结果标签", "较晚结果标签",
    "阶段变化", "客服新增动作",
    "问题是否补救", "补救说明",
    "是否按建议执行", "执行偏差评价",
    "当前最需关注", "风险等级",
    "较早销售触达次数", "新增销售触达次数", "今日销售触达次数",
    "新增客户发言数", "新增客服发言数",
    "_错误", "新增对话",
]

CONFIG_DIR = get_config_dir()
CONFIG_FILE = os.path.join(CONFIG_DIR, "qc_config.json")
CONFIG_EXAMPLE_FILE = os.path.join(CONFIG_DIR, "qc_config.example.json")

DEFAULT_CONCURRENCY = 100

# DeepSeek 账号级并发上限：https://api-docs.deepseek.com/zh-cn/quick_start/rate_limit
MODEL_CONCURRENCY_LIMITS = {
    "deepseek-v4-pro": 500,
    "deepseek-v4-flash": 2500,
}
DEFAULT_MODEL_CONCURRENCY_LIMIT = 500

DEFAULTS = {
    "base_url": "https://api.deepseek.com",
    "model": "deepseek-v4-pro",
    "api_key": "",
    "concurrency": DEFAULT_CONCURRENCY,
    "max_chars": 16000,
    "app_password": "",
    # 前端可选模型列表（管理员可在 qc_config.json 中自定义）
    "models": ["deepseek-v4-pro", "deepseek-chat", "deepseek-reasoner"],
}

CONFIG_KEYS = ("base_url", "model", "api_key", "concurrency", "max_chars", "app_password", "models")
_PLACEHOLDER_KEYS = frozenset({"", "YOUR_API_KEY_HERE"})


def max_concurrency_for_model(model: str | None) -> int:
    """返回当前模型在 DeepSeek 侧的并发上限（用于 UI 校验）。"""
    name = str(model or DEFAULTS.get("model") or "").strip().lower()
    for key, limit in MODEL_CONCURRENCY_LIMITS.items():
        if key in name:
            return limit
    return DEFAULT_MODEL_CONCURRENCY_LIMIT


def default_concurrency_for_model(model: str | None) -> int:
    """推荐默认并发：不超过模型上限，常规账号默认 100。"""
    return min(DEFAULT_CONCURRENCY, max_concurrency_for_model(model))


def mask_api_key(key):
    """脱敏显示 API Key，禁止在界面/日志输出完整 key。"""
    if not key:
        return "（未配置）"
    s = str(key).strip()
    if s in _PLACEHOLDER_KEYS:
        return "（未配置）"
    if len(s) <= 8:
        return "****"
    return f"{s[:4]}****{s[-4:]}"


def redact_secrets(text, cfg=None):
    """从错误信息等文本中移除可能泄露的 API Key。"""
    if text is None:
        return ""
    out = str(text)
    if cfg and cfg.get("api_key"):
        key = str(cfg["api_key"])
        if key and key not in _PLACEHOLDER_KEYS:
            out = out.replace(key, mask_api_key(key))
    return out


def _coerce_config_types(cfg):
    try:
        cfg["concurrency"] = int(cfg.get("concurrency", DEFAULTS["concurrency"]))
    except (TypeError, ValueError):
        cfg["concurrency"] = DEFAULTS["concurrency"]
    try:
        cfg["max_chars"] = int(cfg.get("max_chars", DEFAULTS["max_chars"]))
    except (TypeError, ValueError):
        cfg["max_chars"] = DEFAULTS["max_chars"]
    cfg["app_password"] = str(cfg.get("app_password") or "")

    models = cfg.get("models")
    if isinstance(models, str):
        models = [m.strip() for m in models.split(",") if m.strip()]
    if not isinstance(models, (list, tuple)) or not models:
        models = list(DEFAULTS["models"])
    seen = set()
    norm_models = []
    for m in models:
        m = str(m).strip()
        if m and m not in seen:
            seen.add(m)
            norm_models.append(m)
    if cfg.get("model") and cfg["model"] not in seen:
        norm_models.insert(0, cfg["model"])
    cfg["models"] = norm_models
    return cfg


def merge_config(loaded):
    """合并配置字典，返回 (cfg, error_code)。

    error_code: None=成功, 'missing_api_key', 'invalid_format'
    """
    if not isinstance(loaded, dict):
        return dict(DEFAULTS), "invalid_format"
    cfg = dict(DEFAULTS)
    for k in CONFIG_KEYS:
        if k in loaded and loaded[k] is not None:
            cfg[k] = loaded[k]
    cfg = _coerce_config_types(cfg)
    key = str(cfg.get("api_key", "")).strip()
    if key in _PLACEHOLDER_KEYS:
        return cfg, "missing_api_key"
    return cfg, None


def _admin_setup_hint():
    example = os.path.basename(CONFIG_EXAMPLE_FILE)
    return (
        f"1. 打开文件夹：`{CONFIG_DIR}`\n"
        f"2. 复制 `{example}` 为 `qc_config.json`\n"
        f"3. 填入 `api_key`（及其他项）后保存\n"
        f"4. 刷新本页面\n\n"
        f"或在 Streamlit 部署环境配置 `.streamlit/secrets.toml`（仅管理员操作）。\n"
        f"**请勿将真实 API Key 提交到 Git 或发给使用者。**"
    )


def _read_config_file():
    """读取 qc_config.json，返回 (cfg, error_msg)。"""
    example_name = os.path.basename(CONFIG_EXAMPLE_FILE)
    if not os.path.exists(CONFIG_FILE):
        return dict(DEFAULTS), None  # 文件不存在不算读取出错，由上层决定回退 secrets

    try:
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            loaded = json.load(f)
    except json.JSONDecodeError as e:
        return dict(DEFAULTS), (
            f"配置文件 JSON 格式错误：\n{CONFIG_FILE}\n\n"
            f"错误：{e}\n\n可参考「{example_name}」重新创建。"
        )
    except OSError as e:
        return dict(DEFAULTS), f"无法读取配置文件：\n{CONFIG_FILE}\n\n错误：{e}"

    cfg, code = merge_config(loaded)
    if code == "invalid_format":
        return cfg, (
            f"配置文件格式错误（应为 JSON 对象）：\n{CONFIG_FILE}\n\n"
            f"可参考「{example_name}」重新创建。"
        )
    if code == "missing_api_key":
        return cfg, (
            f"配置文件存在但未填写有效 API Key：\n{CONFIG_FILE}\n\n"
            f"请填入 api_key 后保存。"
        )
    return cfg, None


def _extract_from_secrets(secrets):
    """从 Streamlit secrets 提取配置（支持顶层键或 [qc] 段）。"""
    if secrets is None:
        return None
    try:
        if hasattr(secrets, "get"):
            if secrets.get("qc"):
                return dict(secrets["qc"])
            data = {k: secrets[k] for k in CONFIG_KEYS if k in secrets}
            return data if data else None
    except Exception:
        return None
    return None


def load_config_from_disk():
    """桌面版：仅从 qc_config.json 读取。返回 (cfg, warning_msg)。"""
    cfg, err = _read_config_file()
    if err:
        return cfg, err
    if not os.path.exists(CONFIG_FILE):
        msg = (
            f"未找到配置文件：\n{CONFIG_FILE}\n\n"
            f"请复制同目录下的「{os.path.basename(CONFIG_EXAMPLE_FILE)}」为 qc_config.json，\n"
            f"填入 API Key 后保存，再重新启动程序。"
        )
        return cfg, msg
    return cfg, None


def load_config_for_streamlit(secrets=None):
    """网页版：优先 qc_config.json，其次 st.secrets。

    返回 (cfg, error_msg, source)。
    source: 'file' | 'secrets' | None
    """
    if os.path.exists(CONFIG_FILE):
        cfg, err = _read_config_file()
        if err:
            return cfg, err, None
        return cfg, None, "file"

    loaded = _extract_from_secrets(secrets)
    if loaded:
        cfg, code = merge_config(loaded)
        if code == "missing_api_key":
            return cfg, "管理员未配置 API Key\n\n" + _admin_setup_hint(), None
        if code == "invalid_format":
            return cfg, "Streamlit secrets 配置格式错误。", None
        return cfg, None, "secrets"

    return dict(DEFAULTS), "管理员未配置 API Key\n\n" + _admin_setup_hint(), None


def classify_tier(cust_count):
    if cust_count < TIER_SKIP_MAX:
        return "skip"
    if cust_count <= TIER_LIGHT_MAX:
        return "light"
    return "full"


def _is_redline_hit(value):
    red = str(value or "").strip()
    return bool(red) and red not in ("无", "—", "-", "none", "None")


def _is_high_risk_level(value):
    return str(value or "").strip() in ("高", "高风险")


def _needs_manual_review(value):
    return str(value or "").strip() in ("是", "需要", "需复核", "Y", "y", "yes", "true", "True")


def is_high_risk(row):
    """高风险：红线非空 / 不合格 / 风险等级=高 / 需人工复核 / API 或解析失败。"""
    if row.get("_错误"):
        return True
    if _is_redline_hit(row.get("红线")):
        return True
    if str(row.get("是否合格", "")).strip() == "不合格":
        return True
    if _is_high_risk_level(row.get("风险等级")):
        return True
    if _needs_manual_review(row.get("是否需要人工复核")):
        return True
    return False


def _is_main_problem(value):
    s = str(value or "").strip()
    return bool(s) and s not in ("无明显问题", "无", "—", "-", "none", "None", "无问题")


def _is_concern_issue(value):
    s = str(value or "").strip()
    if not s or s in ("—", "-", "未出现"):
        return False
    return s.startswith("✗") or "✗" in s or "不当" in s


def problem_columns_for_row(row):
    """返回优质客户分析中需要标红的列名（仅问题单元格）。"""
    cols = []
    if row.get("_错误"):
        cols.append("_错误")
    if _is_redline_hit(row.get("红线")):
        cols.append("红线")
    if str(row.get("是否合格", "")).strip() == "不合格":
        cols.append("是否合格")
    if _is_high_risk_level(row.get("风险等级")):
        cols.append("风险等级")
    if _needs_manual_review(row.get("是否需要人工复核")):
        cols.append("是否需要人工复核")
    if _is_main_problem(row.get("主要问题")):
        cols.append("主要问题")
    for name in ("顾虑_效果", "顾虑_信任", "顾虑_价格", "顾虑_拖延", "顾虑_COD"):
        if _is_concern_issue(row.get(name)):
            cols.append(name)
    return cols


def _parse_formatted_ts(value):
    """解析 YYYY-MM-DD HH:MM:SS 格式时间戳。"""
    s = str(value or "").strip()
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d %H:%M:%S")
    except ValueError:
        return None


def _minute_of_day(dt):
    return dt.hour * 60 + dt.minute


def is_low_intent_reply_window(dt):
    """客户首次发言是否落在需统计客服回复超时的时段内。

    以下时段不统计超时：09:30 前、12:00~13:30 午休、17:00~18:30、20:00 后。
    """
    if dt is None:
        return False
    mod = _minute_of_day(dt)
    for (sh, sm), (eh, em) in LOW_INTENT_REPLY_WINDOWS:
        start = sh * 60 + sm
        end = eh * 60 + em
        if start <= mod < end:
            return True
    return False


def problem_columns_for_low_intent_row(row):
    """低意向客户：客户首次发言与客服首次回复间隔超过 10 分钟时标红客服列。

    仅当客户首次发言落在营业统计时段（09:30~12:00、13:30~17:00、18:30~20:00）内才判定。
    """
    cust_ts = _parse_formatted_ts(row.get("客户第一次发言"))
    serv_ts = _parse_formatted_ts(row.get("客服第一次发言"))
    if not cust_ts or not serv_ts:
        return []
    if not is_low_intent_reply_window(cust_ts):
        return []
    gap_min = (serv_ts - cust_ts).total_seconds() / 60
    if gap_min > LOW_INTENT_REPLY_SLOW_MINUTES:
        return ["客服第一次发言"]
    return []


# ============================================================
# 表格解析
# ============================================================
def _norm_col(c):
    return str(c).strip().lower().replace(" ", "").replace("　", "")


def _pick_col(nmap, exacts, contains):
    for cand in exacts:
        for c, n in nmap.items():
            if n == cand:
                return c
    for c, n in nmap.items():
        if any(k in n for k in contains):
            return c
    return None


def detect_columns(cols):
    """智能识别列名（含可选的时间、发送人列）。"""
    nmap = {c: _norm_col(c) for c in cols}
    return {
        "id": _pick_col(nmap, ["会话id", "会话编号", "sessionid", "对话id"], ["会话id", "会话编号", "对话id"]),
        "time": _pick_col(
            nmap,
            ["发言时间", "时间", "发送时间", "消息时间", "timestamp", "日期时间"],
            ["发言时间", "时间", "timestamp", "日期"],
        ),
        "sender": _pick_col(nmap, ["发送人", "说话人", "发言人", "角色", "发送方"], ["发送人", "说话人", "发言人", "角色"]),
        "msg": _pick_col(nmap, ["会话消息", "会话消息内容", "消息内容", "聊天内容", "会话内容",
                               "对话内容", "聊天记录", "消息"], ["消息", "聊天内容", "会话内容", "对话内容"]),
        "contact": _pick_col(nmap, ["联系人", "客户", "客户姓名", "姓名", "客户名"], ["联系人", "客户", "姓名"]),
        "member": _pick_col(nmap, ["接待成员", "客服", "接待人", "坐席", "成员"], ["接待", "客服", "坐席"]),
        "channel": _pick_col(nmap, ["社媒渠道", "渠道", "来源", "社媒"], ["渠道", "社媒", "来源"]),
    }


def suggest_column_options(cols):
    """供 UI 下拉框使用：首项为「（不映射）」。"""
    return ["（不映射）"] + list(cols)


def resolve_column_map(detected, user_overrides=None):
    """合并自动识别与用户手动选择。user_overrides 值为列名或「（不映射）」。"""
    mapping = dict(detected)
    if user_overrides:
        for key, val in user_overrides.items():
            if val and val != "（不映射）":
                mapping[key] = val
            else:
                mapping[key] = None
    return mapping


_EMBEDDED_CHAT_PREFIX_RE = re.compile(r"^\[\d{4}[-/]\d")


def _looks_like_embedded_chat(text):
    """消息是否为「[2026-06-14 14:55:55] 发言人 : 内容」行内时间标签格式。"""
    s = str(text).strip()
    if not s:
        return False
    return bool(_EMBEDDED_CHAT_PREFIX_RE.match(s.splitlines()[0].strip()))


def _extract_embedded_chat_lines(raw, window_start=None, window_end=None):
    """从会话消息文本中按行内 [发言时间] 标签提取发言行。"""
    kept = []
    for part in str(raw).splitlines():
        part = part.strip()
        if not part:
            continue
        ts = _parse_line_timestamp(part)
        if ts is None:
            continue
        if window_start is not None and window_end is not None:
            if not (window_start <= ts < window_end):
                continue
        kept.append(part)
    return kept


def _format_message_line(row, cols):
    """将一行转为对话文本。支持整段消息列，或 时间+发送人+消息 分列格式。"""
    msg_col = cols.get("msg")
    if not msg_col:
        return ""
    raw = str(row.get(msg_col, "")).strip()
    if not raw:
        return ""
    # 已是「[时间] 发言人 : 内容」多行聊天记录，原样保留
    if _looks_like_embedded_chat(raw):
        return raw
    time_col = cols.get("time")
    sender_col = cols.get("sender")
    if time_col or sender_col:
        t = str(row.get(time_col, "")).strip() if time_col else ""
        s = str(row.get(sender_col, "")).strip() if sender_col else "未知"
        m = raw
        if not m:
            return ""
        if t:
            return f'[{t}] {s} : "{m}"'
        return f'{s} : "{m}"'
    return raw


def _content_lines_in_window(row, cols, window_start, window_end):
    """按每条发言的发言时间筛选本行产生的对话行（左闭右开）。"""
    msg_col = cols.get("msg")
    if not msg_col:
        return []
    raw = str(row.get(msg_col, "")).strip()
    if not raw:
        return []

    # 优先：消息内行内时间标签 [2026-06-14 14:55:55] 客户 : "..."
    if _looks_like_embedded_chat(raw):
        return _extract_embedded_chat_lines(raw, window_start, window_end)

    time_col = cols.get("time")
    # 分列格式：以「发言时间」列为准，一行一条发言
    if time_col:
        ts = _parse_datetime_loose(str(row.get(time_col, "")).strip())
        if ts is None:
            return []
        if not (window_start <= ts < window_end):
            return []
        line = _format_message_line(row, cols)
        return [line] if line.strip() else []

    ts = _parse_line_timestamp(raw)
    if ts is None:
        return []
    if window_start <= ts < window_end:
        return [raw]
    return []


def _sessions_from_dataframe(df, filename, cols, time_window=None):
    """从单个 DataFrame 按列映射拆通会话。"""
    sessions = []
    if not cols.get("msg"):
        return sessions, "未映射消息内容列"

    df = df.fillna("")
    cur = None
    has_id = bool(cols.get("id"))

    for ridx, row in df.iterrows():
        if time_window:
            content_lines = _content_lines_in_window(row, cols, time_window[0], time_window[1])
        else:
            line = _format_message_line(row, cols)
            content_lines = [line] if line.strip() else []

        for content in content_lines:
            if not content.strip():
                continue
            sid = str(row.get(cols["id"], "")).strip() if has_id else ""

            def make(_sid):
                return {
                    "会话ID": _sid,
                    "联系人": str(row.get(cols["contact"], "")).strip() if cols.get("contact") else "",
                    "接待成员": str(row.get(cols["member"], "")).strip() if cols.get("member") else "",
                    "渠道": str(row.get(cols["channel"], "")).strip() if cols.get("channel") else "",
                    "对话": content,
                }

            if has_id:
                if sid:
                    if cur is not None and cur["会话ID"] == sid:
                        cur["对话"] += "\n" + content
                    else:
                        cur = make(sid)
                        sessions.append(cur)
                elif cur is not None:
                    cur["对话"] += "\n" + content
            else:
                sessions.append(make(f"{filename}#{ridx}"))

    return sessions, None


def load_sessions(files=None, file_dfs=None, column_map=None, time_scope=TIME_SCOPE_ALL, now=None):
    """读取会话。支持文件路径列表或 (name, DataFrame) 列表。

    返回 (sessions, diag, window_info)。
    window_info: 当天模式时为窗口说明，否则为 None。
    """
    sessions = []
    diag = []
    sources = []
    time_window, window_info = resolve_time_scope_window(time_scope, now)

    if file_dfs:
        for name, df in file_dfs:
            sources.append((name, df))
    elif files:
        for f in files:
            name = os.path.basename(f)
            try:
                df = pd.read_excel(f, sheet_name=0, dtype=str)
                sources.append((name, df))
            except Exception as e:
                diag.append({"file": name, "error": str(e), "columns": [], "rows": 0})
    else:
        return [], diag

    for name, df in sources:
        auto = detect_columns(df.columns)
        cols = resolve_column_map(auto, column_map) if column_map else auto
        diag.append({
            "file": name,
            "columns": list(df.columns),
            "detected": cols,
            "rows": len(df),
        })
        part, err = _sessions_from_dataframe(df, name, cols, time_window=None)
        if err and not part:
            diag[-1]["error"] = err
            continue
        if time_window:
            start, end = time_window
            scoped_part = []
            for s in part:
                full = s["对话"]
                dialog = filter_dialog_by_window(full, start, end)
                if not dialog.strip():
                    continue
                item = dict(s)
                item["对话_全量"] = full
                item["对话"] = dialog
                scoped_part.append(item)
            part = scoped_part
        sessions.extend(part)

    uniq = {}
    for s in sessions:
        sid = s["会话ID"]
        if sid not in uniq or len(s["对话"]) > len(uniq[sid]["对话"]):
            uniq[sid] = s
    out = []
    for s in uniq.values():
        contact = s.get("联系人", "")
        if contact:
            s = dict(s)
            s["对话"] = normalize_customer_speaker_in_dialog(s.get("对话") or "", contact)
        out.append(s)
    return out, diag, window_info


_SYSTEM_ROLES = frozenset({"自动化", "自定义", "其他平台"})

# 客户/客服发言数规则（对齐 参考代码/new7.2.py，按行解析 `[时间] 发言人 : 内容`）
CUSTOMER_SEG_RE = re.compile(r'客户\s*:\s*[“"\u201c]?\s*(.*?)[”"\u201d]?(?=$|\r|\n)')
AGENT_SEG_RE = re.compile(r'客服\s*:\s*[“"\u201c]?\s*(.*?)[”"\u201d]?(?=$|\r|\n)')
IMG_TOKEN = "图片"
EMOJI_POWER_RE = re.compile(r'^【\s*(?:💪\s*)?\d+\+\s*】$')
# 宽容匹配【💪40+】类力度标签（兼容不同括号/emoji 编码）
POWER_TAG_LOOSE_RE = re.compile(r'^【[^】]*\d+\+[^】]*】$')


def _is_power_emoji_tag(seg):
    s = (seg or "").strip()
    return bool(EMOJI_POWER_RE.match(s) or POWER_TAG_LOOSE_RE.match(s))


def _extract_message_content_from_line(line):
    """从单行对话提取冒号后的消息正文。"""
    line = (line or "").strip()
    if not line:
        return None
    rest = line
    if line.startswith("["):
        m = re.match(r"^\[[^\]]+\]\s*(.+)$", line)
        if not m:
            return None
        rest = m.group(1)
    m = re.match(r"^(.+?)\s*[:：]\s*(.*)$", rest, re.DOTALL)
    if not m:
        return None
    return _strip_wrapping_quotes(m.group(2))


def _strip_wrapping_quotes(text):
    s = (text or "").strip()
    pairs = (
        ("\u201c", "\u201d"),
        ("\u2018", "\u2019"),
        ('"', '"'),
        ('"', '"'),
        ("「", "」"),
        ("『", "』"),
    )
    for left, right in pairs:
        if s.startswith(left) and s.endswith(right) and len(s) > len(left) + len(right) - 1:
            return s[len(left):-len(right)].strip()
    return s


def _iter_dialog_lines(text):
    for line in str(text).splitlines():
        line = line.strip()
        if line:
            yield line


def extract_customer_segments(text):
    """提取所有客户发言正文（优先按行解析，兼容整段文本正则）。"""
    if not isinstance(text, str) or not text.strip():
        return []
    segs = []
    for line in _iter_dialog_lines(text):
        if _parse_speaker(line) != "客户":
            continue
        content = _extract_message_content_from_line(line)
        if content is not None:
            segs.append(content)
    if segs:
        return segs
    return [(seg or "").strip() for seg in CUSTOMER_SEG_RE.findall(text)]


def extract_agent_segments(text):
    """提取客服/销售发言正文（具名销售如杜凯、王质一，及「客服 :」）。"""
    if not isinstance(text, str) or not text.strip():
        return []
    segs = []
    for line in _iter_dialog_lines(text):
        who = _parse_speaker(line)
        if not who or who == "客户" or who in _SYSTEM_ROLES:
            continue
        content = _extract_message_content_from_line(line)
        if content is not None:
            segs.append(content)
    if segs:
        return segs
    return [(seg or "").strip() for seg in AGENT_SEG_RE.findall(text)]


def _is_invalid_customer_segment(seg):
    s = (seg or "").strip()
    return not s or s == IMG_TOKEN or s == "1"


def _is_invalid_agent_segment(seg):
    s = (seg or "").strip()
    return not s or s == "1" or _is_power_emoji_tag(s)


def count_valid_customer_segments(text):
    """客户发言数：非图片 / 非孤立「1」的客户发言条数（含【50+】等广告进线力度标签）。"""
    return sum(1 for seg in extract_customer_segments(text) if not _is_invalid_customer_segment(seg))


def count_valid_agent_segments(text):
    """客服发言数：排除表情与孤立「1」；图片/媒体类消息计入。"""
    return sum(1 for seg in extract_agent_segments(text) if not _is_invalid_agent_segment(seg))


def count_roles(text):
    """统计有效客户/客服发言条数（用于分档与报告）。"""
    return count_valid_customer_segments(text), count_valid_agent_segments(text)


def _parse_speaker(line):
    """从对话行解析说话人。支持带/不带时间戳，中英文冒号。"""
    line = line.strip()
    if not line:
        return None
    rest = line
    if line.startswith("["):
        m = re.match(r"^\[[^\]]+\]\s*(.+)$", line)
        if m:
            rest = m.group(1)
        else:
            return None
    m = re.match(r"^(.+?)\s*[:：]\s*", rest)
    if not m:
        return None
    return m.group(1).strip()


def normalize_customer_speaker_in_dialog(text, contact_name):
    """将「联系人姓名 : …」统一为「客户 : …」，兼容部分导出把访客名写在发言人列的情况。"""
    contact = (contact_name or "").strip()
    if not contact or not str(text or "").strip():
        return text
    escaped = re.escape(contact)
    pattern = re.compile(
        rf"^(\[[^\]]+\]\s*)?{escaped}(\s*[:：])",
        re.MULTILINE | re.IGNORECASE,
    )
    return pattern.sub(r"\1客户\2", str(text))


# ============================================================
# 发言时间解析
# ============================================================
_DATETIME_FORMATS = (
    "%Y-%m-%d %H:%M:%S",
    "%Y/%m/%d %H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%Y/%m/%d %H:%M",
    "%m-%d %H:%M:%S",
    "%m/%d %H:%M:%S",
    "%m-%d %H:%M",
    "%m/%d %H:%M",
)


def _parse_datetime_loose(text):
    s = str(text).strip()
    if not s:
        return None
    for fmt in _DATETIME_FORMATS:
        try:
            dt = datetime.strptime(s, fmt)
            if fmt.startswith("%m"):
                dt = dt.replace(year=datetime.now().year)
            return dt
        except ValueError:
            continue
    try:
        parsed = pd.to_datetime(s, errors="coerce")
        if pd.isna(parsed):
            return None
        return parsed.to_pydatetime()
    except Exception:
        return None


def _parse_line_timestamp(line):
    """从对话行解析发言时间，如 [2026-01-01 10:00:00]。"""
    line = line.strip()
    if not line.startswith("["):
        return None
    m = re.match(r"^\[([^\]]+)\]", line)
    if not m:
        return None
    return _parse_datetime_loose(m.group(1).strip())


# ============================================================
# 分析时间范围（当天 = 昨日 20:00 ~ 今日 20:00）
# ============================================================

def get_business_day_window(now=None):
    """当天营业日窗口：[昨日 20:00, 今日 20:00)；若已过今日 20:00 则滚动到下一窗口。"""
    now = now or datetime.now()
    cutoff = now.replace(hour=BUSINESS_DAY_CUTOFF_HOUR, minute=0, second=0, microsecond=0)
    if now >= cutoff:
        start = cutoff
        end = cutoff + timedelta(days=1)
    else:
        end = cutoff
        start = cutoff - timedelta(days=1)
    return start, end


def format_business_day_window(start, end):
    return f"{start.strftime('%Y-%m-%d %H:%M')} ~ {end.strftime('%Y-%m-%d %H:%M')}"


def resolve_time_scope_window(scope=TIME_SCOPE_ALL, now=None):
    """解析分析范围，返回 (time_window, window_info)。"""
    if scope in (TIME_SCOPE_ALL, TIME_SCOPE_LABELS[TIME_SCOPE_ALL], "全部"):
        return None, None
    start, end = get_business_day_window(now)
    info = {"start": start, "end": end, "label": format_business_day_window(start, end)}
    return (start, end), info


def filter_dialog_by_window(text, window_start, window_end):
    """仅保留发言时间落在窗口内的对话行（左闭右开）。"""
    if _looks_like_embedded_chat(text):
        return "\n".join(_extract_embedded_chat_lines(text, window_start, window_end))
    kept = []
    for line in text.splitlines():
        if not line.strip():
            continue
        ts = _parse_line_timestamp(line)
        if ts is None:
            continue
        if window_start <= ts < window_end:
            kept.append(line)
    return "\n".join(kept)


def filter_sessions_by_window(sessions, start, end):
    """按任意自定义时间窗口 [start, end) 过滤会话的对话行。

    返回 (filtered_sessions, window_info)。
    """
    window_info = {
        "start": start,
        "end": end,
        "label": format_business_day_window(start, end),
    }
    filtered = []
    for session in sessions:
        dialog = filter_dialog_by_window(session["对话"], start, end)
        if not dialog.strip():
            continue
        scoped = dict(session)
        scoped["对话"] = dialog
        scoped["对话_全量"] = session.get("对话_全量", session["对话"])
        filtered.append(scoped)
    return filtered, window_info


def apply_time_scope_to_sessions(sessions, scope=TIME_SCOPE_ALL, now=None):
    """按发言时间过滤已加载会话（兜底）。优先在 load_sessions 阶段过滤。"""
    time_window, window_info = resolve_time_scope_window(scope, now)
    if time_window is None:
        return list(sessions), None

    start, end = time_window
    filtered = []
    for session in sessions:
        dialog = filter_dialog_by_window(session["对话"], start, end)
        if not dialog.strip():
            continue
        scoped = dict(session)
        scoped["对话"] = dialog
        scoped["对话_全量"] = session.get("对话_全量", session["对话"])
        filtered.append(scoped)
    return filtered, window_info


# ============================================================
# 低意向客户 · 发言频率与间隔（纯 Python，不调 API）
# ============================================================

def _is_service_speaker(who):
    return who is not None and who != "客户" and who not in _SYSTEM_ROLES


def _extract_timestamps_by_speaker(text, speaker):
    """speaker: 'customer' | 'service'；仅统计有效发言片段的时间。"""
    stamps = []
    for line in text.splitlines():
        who = _parse_speaker(line)
        if speaker == "customer" and who != "客户":
            continue
        if speaker == "service" and not _is_service_speaker(who):
            continue
        content = _extract_message_content_from_line(line)
        if speaker == "customer" and _is_invalid_customer_segment(content):
            continue
        if speaker == "service" and _is_invalid_agent_segment(content):
            continue
        ts = _parse_line_timestamp(line)
        if ts is not None:
            stamps.append(ts)
    return stamps


def _extract_all_customer_timestamps(text):
    """客户每条发言的时间戳（含表情/图片等，按对话顺序）。"""
    stamps = []
    for line in _iter_dialog_lines(text):
        if _parse_speaker(line) != "客户":
            continue
        ts = _parse_line_timestamp(line)
        if ts is not None:
            stamps.append(ts)
    return stamps


def _extract_all_service_timestamps(text):
    """客服/销售每条回复的时间戳（按对话顺序）。"""
    stamps = []
    for line in _iter_dialog_lines(text):
        who = _parse_speaker(line)
        if not who or who == "客户" or who in _SYSTEM_ROLES:
            continue
        ts = _parse_line_timestamp(line)
        if ts is not None:
            stamps.append(ts)
    return stamps


def _extract_customer_timestamps(text):
    return _extract_timestamps_by_speaker(text, "customer")


def _last_customer_message_timestamp(text):
    """客户最后一条消息时间：优先有效发言，否则取无效发言（表情/图片/1）中最后一条。"""
    valid = _extract_customer_timestamps(text)
    if valid:
        return valid[-1]
    all_cust = _extract_all_customer_timestamps(text)
    return all_cust[-1] if all_cust else None


def _extract_service_timestamps(text):
    return _extract_timestamps_by_speaker(text, "service")


def _extract_inbound_timestamp(text):
    """进线时间：会话中第一条可解析时间的消息（客户或客服）。"""
    for line in text.splitlines():
        who = _parse_speaker(line)
        if who is None or who in _SYSTEM_ROLES:
            continue
        ts = _parse_line_timestamp(line)
        if ts is not None:
            return ts
    return None


def cluster_customer_utterances(timestamps, gap_minutes=UTTERANCE_GAP_MINUTES):
    """5 分钟内连续客户消息合并为一次有效发言。"""
    if not timestamps:
        return []
    gap_sec = gap_minutes * 60
    bursts = [[timestamps[0]]]
    for ts in timestamps[1:]:
        if (ts - bursts[-1][-1]).total_seconds() <= gap_sec:
            bursts[-1].append(ts)
        else:
            bursts.append([ts])
    return bursts


def _minutes_between(a, b):
    return round((b - a).total_seconds() / 60, 1)


def _format_duration_minutes(minutes):
    if minutes is None or minutes == "":
        return ""
    try:
        m = float(minutes)
    except (TypeError, ValueError):
        return ""
    if m < 60:
        return f"{m:.1f}分钟"
    if m < 24 * 60:
        return f"{m / 60:.1f}小时"
    return f"{m / (24 * 60):.1f}天"


def _frequency_label(burst_count, span_minutes):
    """将有效次数与会话跨度格式化为可读频率。"""
    if burst_count <= 0:
        return "无"
    if span_minutes is None or span_minutes == "":
        return f"{burst_count}次"
    try:
        span = float(span_minutes)
    except (TypeError, ValueError):
        return f"{burst_count}次"
    if span <= 0:
        return f"{burst_count}次"
    return f"{burst_count}次/{_format_duration_minutes(span)}"


def _burst_intervals(bursts):
    """相邻有效发言/跟进之间的间隔（分钟）。"""
    if len(bursts) < 2:
        return []
    starts = [b[0] for b in bursts]
    return [_minutes_between(starts[i], starts[i + 1]) for i in range(len(starts) - 1)]


def _format_ts(dt):
    return dt.strftime("%Y-%m-%d %H:%M:%S") if dt else ""


def _service_burst_starts(serv_timestamps, gap_minutes=UTTERANCE_GAP_MINUTES):
    return [b[0] for b in cluster_customer_utterances(serv_timestamps, gap_minutes=gap_minutes)]


def count_service_touch_bursts(text, gap_minutes=PROGRESS_COMPARE_TOUCH_GAP_MINUTES):
    """销售触达次数：有效客服发言按 gap_minutes 内连续多条合并为 1 次（进度对比默认 10 分钟）。"""
    stamps = _extract_service_timestamps(text)
    if not stamps:
        return 0
    return len(_service_burst_starts(stamps, gap_minutes=gap_minutes))


def _parse_dialog_events(text):
    """按时间顺序解析对话行（客户/客服有效发言）。"""
    events = []
    for line in _iter_dialog_lines(text):
        who = _parse_speaker(line)
        ts = _parse_line_timestamp(line)
        content = _extract_message_content_from_line(line)
        if who == "客户":
            if ts is None:
                continue
            events.append({
                "role": "customer",
                "ts": ts,
                "valid": not _is_invalid_customer_segment(content),
            })
        elif _is_service_speaker(who):
            if ts is None:
                continue
            events.append({
                "role": "service",
                "ts": ts,
                "valid": not _is_invalid_agent_segment(content),
            })
    return events


def _build_speech_turns(events, max_turns=3):
    """交替轮次：每轮取客户首条与客服首条（有效发言，不按 5 分钟合并）。"""
    turns = []
    i = 0
    n = len(events)

    while i < n and len(turns) < max_turns:
        while i < n and not (events[i]["role"] == "customer" and events[i]["valid"]):
            i += 1
        if i >= n:
            break

        cust_first = events[i]["ts"]
        i += 1
        while i < n and events[i]["role"] == "customer" and events[i]["valid"]:
            i += 1

        serv_first = None
        if i < n and events[i]["role"] == "service" and events[i]["valid"]:
            serv_first = events[i]["ts"]
            i += 1
            while i < n and events[i]["role"] == "service" and events[i]["valid"]:
                i += 1

        turns.append({"cust": cust_first, "serv": serv_first})

    return turns


def _low_intent_round_fields(text):
    """低意向：客户/客服按对话轮次各记首条时间，不按 5 分钟合并。"""
    events = _parse_dialog_events(text)
    turns = _build_speech_turns(events)
    values = {col: "" for pair in LOW_INTENT_SPEECH_PAIRS for col in pair}

    for i, (cust_col, serv_col) in enumerate(LOW_INTENT_SPEECH_PAIRS):
        if i < len(turns):
            values[cust_col] = _format_ts(turns[i]["cust"])
            if turns[i]["serv"] is not None:
                values[serv_col] = _format_ts(turns[i]["serv"])

    return values


def _low_intent_time_fields(text):
    """客户与客服各次发言时间（基于当前对话文本，已含当天筛选结果）。"""
    return _low_intent_round_fields(text)


def _low_intent_screen_tag(utterance_count, raw_count, intervals):
    if utterance_count <= 0:
        return "无客户发言"
    if utterance_count == 1:
        if raw_count == 1:
            return "单次留言"
        return "单次连发"
    if utterance_count == 2:
        gap = intervals[0] if intervals else None
        if gap is not None and gap <= 30:
            return "二次短回访"
        if gap is not None and gap <= 24 * 60:
            return "二次中回访"
        return "二次长回访"
    return "多次试探"


def _low_intent_screen_note(text):
    """筛选说明：客户沉默时间 + 之后销售发送次数与各次时间（5 分钟合并计一次）。"""
    last_cust = _last_customer_message_timestamp(text)
    serv_stamps = _extract_service_timestamps(text)

    if last_cust is None:
        if extract_customer_segments(text):
            return "客户发言缺少可解析时间，无法统计沉默时间与销售跟进"
        return "对话中未识别到客户发言"

    after_serv = [ts for ts in serv_stamps if ts > last_cust]
    burst_starts = _service_burst_starts(after_serv)
    silence = _format_ts(last_cust)
    only_invalid = not _extract_customer_timestamps(text) and extract_customer_segments(text)
    if only_invalid:
        silence = f"{silence}（无效发言）"
    if not burst_starts:
        cust_count = count_valid_customer_segments(text)
        hint = ""
        if cust_count < TIER_LIGHT_MAX:
            hint = "；当天发言不多，应及时跟进、多发带价值消息"
        return f"客户沉默于{silence}，之后销售未发送{hint}"
    times = "、".join(_format_ts(ts) for ts in burst_starts)
    return f"客户沉默于{silence}，之后销售发送{len(burst_starts)}次：{times}"


def analyze_low_intent_frequency(text):
    """低意向会话：客户首次发言与客服回复时间（对话文本已按当天/全部范围过滤）。"""
    raw_cust, raw_serv = count_roles(text)
    cust_stamps = _extract_customer_timestamps(text)
    serv_stamps = _extract_service_timestamps(text)
    empty = {c: "" for c in LOW_INTENT_ANALYSIS_COLUMNS}
    time_fields = _low_intent_time_fields(text)

    if raw_cust == 0 and raw_serv == 0:
        invalid_cust = len(extract_customer_segments(text)) - raw_cust
        return {
            **empty,
            **time_fields,
            "快速筛选": "0互动" if invalid_cust > 0 else "无有效对话",
            "筛选说明": _low_intent_screen_note(text),
        }

    if not cust_stamps and not serv_stamps:
        note = _low_intent_screen_note(text)
        if not _extract_customer_timestamps(text) and not _extract_service_timestamps(text):
            note = "发言行缺少可解析时间，无法统计客户末条消息与客服跟进"
        return {
            **empty,
            **time_fields,
            "快速筛选": "无时间戳",
            "筛选说明": note,
        }

    cust_bursts = cluster_customer_utterances(cust_stamps) if cust_stamps else []
    utterance_count = len(cust_bursts)
    cust_intervals = _burst_intervals(cust_bursts)

    tag = _low_intent_screen_tag(utterance_count, raw_cust, cust_intervals)
    note = _low_intent_screen_note(text)

    return {
        **time_fields,
        "快速筛选": tag,
        "筛选说明": note,
    }


# ============================================================
# 大模型调用
# ============================================================
def _truncate_dialog_for_llm(text, max_chars):
    """超长对话保留首尾：报价/人设多在前段，最新客户状态在尾段，中间省略。

    旧做法只保留结尾(text[-max_chars:])会丢掉前段的报价，导致误判「未报价」。
    """
    text = text or ""
    if len(text) <= max_chars:
        return text
    lines = text.splitlines()
    head_budget = int(max_chars * 0.6)
    tail_budget = max_chars - head_budget
    n = len(lines)
    hi, h = 0, 0
    while hi < n and h + len(lines[hi]) + 1 <= head_budget:
        h += len(lines[hi]) + 1
        hi += 1
    ti, t = n, 0
    while ti - 1 >= hi and t + len(lines[ti - 1]) + 1 <= tail_budget:
        ti -= 1
        t += len(lines[ti]) + 1
    if ti <= hi:  # 首尾相接，无需省略标记
        return "\n".join(lines[:max(hi, 1)])
    omitted = ti - hi
    return "\n".join(lines[:hi]) + f"\n…（中间省略 {omitted} 行）…\n" + "\n".join(lines[ti:])


def fetch_available_models(cfg, timeout=15):
    """从 OpenAI 兼容接口拉取可用模型列表（GET {base_url}/models）。

    返回 (models, error)：成功时 models 为模型名列表、error 为 None；
    失败时 models 为 []、error 为脱敏后的提示。
    """
    base = str(cfg.get("base_url") or "").rstrip("/")
    key = str(cfg.get("api_key") or "").strip()
    if not base:
        return [], "未配置 base_url"
    if key in _PLACEHOLDER_KEYS:
        return [], "未配置有效 API Key"

    url = base + "/models"
    headers = {"Authorization": "Bearer " + key}
    try:
        r = requests.get(url, headers=headers, timeout=timeout)
        if r.status_code != 200:
            return [], redact_secrets(f"HTTP {r.status_code}: {r.text[:200]}", cfg)
        data = r.json()
    except Exception as e:
        return [], redact_secrets(str(e), cfg)

    items = data.get("data") if isinstance(data, dict) else None
    if not isinstance(items, list):
        return [], "接口返回格式异常（缺少 data 列表）"

    models = []
    seen = set()
    for it in items:
        mid = None
        if isinstance(it, dict):
            mid = it.get("id") or it.get("model")
        elif isinstance(it, str):
            mid = it
        if mid:
            mid = str(mid).strip()
            if mid and mid not in seen:
                seen.add(mid)
                models.append(mid)
    return models, None


def call_llm_prompt(cfg, system_prompt, user_prompt, max_chars=None, truncate=True):
    """通用 LLM 调用（质检 / 成交学习共用），返回解析后的 JSON dict。"""
    limit = max_chars or max(int(cfg.get("max_chars", 16000)), 24000)
    text = (
        user_prompt
        if not truncate
        else _truncate_dialog_for_llm(user_prompt, limit)
    )
    url = cfg["base_url"].rstrip("/") + "/chat/completions"
    headers = {
        "Authorization": "Bearer " + cfg["api_key"],
        "Content-Type": "application/json",
    }
    payload = {
        "model": cfg["model"],
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": text},
        ],
        "temperature": 0.2,
        "response_format": {"type": "json_object"},
        "stream": False,
    }
    last_err = None
    for attempt in range(3):
        try:
            r = requests.post(url, headers=headers, json=payload, timeout=120)
            if r.status_code != 200:
                last_err = f"HTTP {r.status_code}: {r.text[:200]}"
                if "response_format" in r.text:
                    payload.pop("response_format", None)
                time.sleep(1.5 * (attempt + 1))
                continue
            data = r.json()
            content = data["choices"][0]["message"]["content"]
            return parse_json_loose(content)
        except Exception as e:
            last_err = str(e)
            time.sleep(1.5 * (attempt + 1))
    return {"_错误": redact_secrets(last_err or "未知错误", cfg)}


def call_llm(cfg, dialog_text, tier="full", note=None, deal_reference=None):
    if tier == "light":
        system_prompt = enrich_system_prompt(load_prompt("lite_qc.md"), tier)
        max_chars = LITE_MAX_CHARS
        user_tail = "请按标准质检以下整通会话，只输出 JSON：\n\n"
    elif tier == "deep_review":
        system_prompt = load_prompt(DEEP_REVIEW_PROMPT_FILE)
        max_chars = max(int(cfg.get("max_chars", 16000)), 24000)
        user_tail = "请对以下整通会话做【深度复盘】，尽可能全面分析客户卡点与下一步动作，只输出 JSON：\n\n"
    elif tier == "progress_compare":
        system_prompt = load_prompt(PROGRESS_COMPARE_PROMPT_FILE)
        max_chars = max(int(cfg.get("max_chars", 16000)), 24000)
        # 调用方已把两次快照与新增对话拼好为完整载荷，这里不再额外加引导句。
        user_tail = ""
    else:
        system_prompt = enrich_system_prompt(load_prompt("full_qc.md"), tier)
        max_chars = cfg["max_chars"]
        user_tail = "请按标准质检以下整通会话，只输出 JSON：\n\n"

    text = _truncate_dialog_for_llm(dialog_text, max_chars)
    prefix_parts = []
    if deal_reference:
        prefix_parts.append(deal_reference.strip())
    if note:
        prefix_parts.append(note.strip())
    prefix = "\n\n".join(prefix_parts)
    user_msg = (prefix + "\n\n" if prefix else "") + user_tail + text
    url = cfg["base_url"].rstrip("/") + "/chat/completions"
    headers = {
        "Authorization": "Bearer " + cfg["api_key"],
        "Content-Type": "application/json",
    }
    payload = {
        "model": cfg["model"],
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_msg},
        ],
        "temperature": 0.2,
        "response_format": {"type": "json_object"},
        "stream": False,
    }
    last_err = None
    for attempt in range(3):
        try:
            r = requests.post(url, headers=headers, json=payload, timeout=120)
            if r.status_code != 200:
                last_err = f"HTTP {r.status_code}: {r.text[:200]}"
                if "response_format" in r.text:
                    payload.pop("response_format", None)
                time.sleep(1.5 * (attempt + 1))
                continue
            data = r.json()
            content = data["choices"][0]["message"]["content"]
            return parse_json_loose(content)
        except Exception as e:
            last_err = str(e)
            time.sleep(1.5 * (attempt + 1))
    return {"_错误": redact_secrets(last_err or "未知错误", cfg)}


def parse_json_loose(s):
    s = s.strip()
    if s.startswith("```"):
        s = s.strip("`")
        if s.lower().startswith("json"):
            s = s[4:]
    try:
        return json.loads(s)
    except Exception:
        a, b = s.find("{"), s.rfind("}")
        if a != -1 and b != -1 and b > a:
            try:
                return json.loads(s[a:b + 1])
            except Exception:
                pass
    return {"_错误": "返回非JSON", "原始返回": s[:500]}


# ============================================================
# 批量质检（桌面 / 网页共用）
# ============================================================
_MISSING_OPENING_HINTS = (
    "未执行", "未进行", "未完成标准流程", "未完成流程", "人设", "症状菜单",
    "未引入产品", "未问诊", "开头", "缺失", "全程闲聊", "闲聊", "失职",
    "直接跳", "SOP", "未铺垫", "省略", "流程中断",
)


def _suspect_missing_opening(res):
    """首轮结果是否疑似「开头缺失/未执行SOP/老客户建联被误判」——需结合完整历史复核。"""
    if not isinstance(res, dict):
        return False
    blob = " ".join(str(res.get(k, "")) for k in ("主要问题", "红线", "改进建议"))
    return any(h in blob for h in _MISSING_OPENING_HINTS)


_FULL_HISTORY_NOTE = (
    "⚠️ 以下是该客户的【完整历史对话】（含所选日期之前的记录）。请据此重新判断：\n"
    "若开头的人设/问诊/方案价值是在更早的历史里已经做过、当前只是老客户日常建联/闲聊"
    "（典型特征：存在大量更早日期的记录、客户与客服发言比约 1:2~1:3、内容偏问候闲聊或加快节奏），"
    "则属【正确行为】，不要再判「未执行SOP / 开头缺失 / 失职 / 流程中断」。"
)


def _resolve_deal_context(cfg, dialog: str, use_deal_context: bool | None) -> tuple[str, list, str, str]:
    """返回 (reference_note, deals, count_label, summary)。"""
    enabled = cfg.get("deal_context_enabled", True) if use_deal_context is None else bool(use_deal_context)
    if not enabled:
        return "", [], "", ""
    try:
        from deal_context import get_deal_context_for_qc, summarize_deal_refs

        note, deals = get_deal_context_for_qc(dialog, cfg)
        summary = summarize_deal_refs(deals)
        count = str(len(deals)) if deals else ""
        return note, deals, count, summary
    except Exception as e:
        logger.warning("成交经验检索失败: %s", e)
        return "", [], "", ""


def qc_one_session(cfg, session, *, use_deal_context=None):
    """质检单通会话，异常不向外抛出。"""
    cust, serv = count_roles(session["对话"])
    tier = classify_tier(cust)
    row = {
        "会话ID": session["会话ID"],
        "联系人": session["联系人"],
        "接待成员": session["接待成员"],
        "渠道": session["渠道"],
        "客户发言数": cust,
        "客服发言数": serv,
        "质检档位": TIER_LABEL[tier],
        "原始对话": session["对话"],
        "成交参考案例数": "",
        "成交参考摘要": "",
    }
    if tier == "skip":
        row["结果标签"] = "规则筛选"
        row.update(analyze_low_intent_frequency(session["对话"]))
        return row, tier
    try:
        deal_ref, _deals, ref_count, ref_summary = _resolve_deal_context(
            cfg, session["对话"], use_deal_context
        )
        row["成交参考案例数"] = ref_count
        row["成交参考摘要"] = ref_summary

        res = call_llm(cfg, session["对话"], tier=tier, deal_reference=deal_ref or None)
        full_dialog = session.get("对话_全量")
        if (full_dialog and full_dialog.strip() != (session["对话"] or "").strip()
                and _suspect_missing_opening(res)):
            full_ref, _, full_count, full_summary = _resolve_deal_context(
                cfg, full_dialog, use_deal_context
            )
            res_full = call_llm(
                cfg, full_dialog, tier="full",
                note=_FULL_HISTORY_NOTE,
                deal_reference=full_ref or None,
            )
            if isinstance(res_full, dict) and "_错误" not in res_full:
                res = res_full
                row["成交参考案例数"] = full_count or ref_count
                row["成交参考摘要"] = full_summary or ref_summary
                res["改进建议"] = ("【已结合完整历史复核】 " + str(res.get("改进建议", ""))).strip()
        for k, v in res.items():
            val = v if not isinstance(v, (dict, list)) else json.dumps(v, ensure_ascii=False)
            if isinstance(val, str):
                val = redact_secrets(val, cfg)
            row[k] = val
    except Exception as e:
        row["_错误"] = redact_secrets(str(e), cfg)
    return row, tier


def compare_qc_deal_context(cfg, session) -> dict:
    """效果验证：同一通会话分别做「有/无成交经验参考」质检并对比差异。"""
    row_with, tier = qc_one_session(cfg, session, use_deal_context=True)
    row_without, _ = qc_one_session(cfg, session, use_deal_context=False)
    from deal_context import diff_qc_rows

    return {
        "tier": tier,
        "with_context": row_with,
        "without_context": row_without,
        "diffs": diff_qc_rows(row_with, row_without),
        "deal_ref_summary": row_with.get("成交参考摘要", ""),
    }


def get_session_full_dialog(session):
    """获取会话完整聊天记录（不受当天/自定义时间窗口截断）。"""
    full = session.get("对话_全量")
    if full and str(full).strip():
        return str(full)
    return str(session.get("对话", "") or "")


_DEEP_REVIEW_FULL_NOTE = (
    "以下为该客户的【完整聊天记录】（含时间筛选范围外的全部历史发言）。"
    "请基于完整上下文做深度复盘，不要只根据片段下结论。"
)


def qualifies_for_deep_review(session):
    """客户发言数是否超过深度复盘阈值（> DEEP_REVIEW_MIN_CUSTOMER_SPEECHES）。"""
    cust, _ = count_roles(get_session_full_dialog(session))
    try:
        return int(cust) > DEEP_REVIEW_MIN_CUSTOMER_SPEECHES
    except (TypeError, ValueError):
        return False


def deep_review_one_session(cfg, session):
    """对单通会话做深度复盘（独立任务，默认使用完整聊天记录）。"""
    dialog = get_session_full_dialog(session)
    cust, serv = count_roles(dialog)
    row = {
        "会话ID": session.get("会话ID", ""),
        "联系人": session.get("联系人", ""),
        "接待成员": session.get("接待成员", ""),
        "渠道": session.get("渠道", ""),
        "客户发言数": cust,
        "客服发言数": serv,
        "原始对话": dialog,
    }
    note = _DEEP_REVIEW_FULL_NOTE if session.get("对话_全量") else None
    try:
        res = call_llm(cfg, dialog, tier="deep_review", note=note)
        for k, v in res.items():
            val = v if not isinstance(v, (dict, list)) else json.dumps(v, ensure_ascii=False)
            if isinstance(val, str):
                val = redact_secrets(val, cfg)
            row[k] = val
    except Exception as e:
        row["_错误"] = redact_secrets(str(e), cfg)
    return row


def _group_deep_review_by_note(eligible):
    """深度复盘按 user 前缀分组：无完整历史 note 先跑，有 note 后跑（利于缓存）。"""
    without_note, with_note = [], []
    for idx, s in enumerate(eligible):
        if s.get("对话_全量"):
            with_note.append((idx, s))
        else:
            without_note.append((idx, s))
    waves = []
    if without_note:
        waves.append(without_note)
    if with_note:
        waves.append(with_note)
    return waves


def run_deep_review_batch(sessions, cfg, on_progress=None, timing_out=None):
    """在常规质检完成后，对符合条件的会话并发执行深度复盘（按 note 前缀分两批）。"""
    eligible = [s for s in sessions if qualifies_for_deep_review(s)]
    results = [None] * len(eligible)
    if not eligible:
        if timing_out is not None:
            timing_out.clear()
        return results

    done = 0
    total = len(eligible)
    conc = max(1, int(cfg.get("concurrency", DEFAULT_CONCURRENCY)))
    started_at = datetime.now()
    batch_t0 = time.perf_counter()
    total_seconds = 0.0

    for wave in _group_deep_review_by_note(eligible):
        def task(entry):
            i, s = entry
            t0 = time.perf_counter()
            row = deep_review_one_session(cfg, s)
            return i, row, time.perf_counter() - t0

        for i, row, elapsed in _run_parallel_with_cache_warmup(wave, conc, task, cache_warmup=True):
            results[i] = row
            total_seconds += elapsed
            done += 1
            if on_progress:
                on_progress(done, total, row)

    if timing_out is not None:
        timing_out.clear()
        timing_out.update({
            "started_at": started_at,
            "ended_at": datetime.now(),
            "wall_seconds": time.perf_counter() - batch_t0,
            "concurrency": conc,
            "count": total,
            "total_seconds": total_seconds,
        })
    return results


def format_duration(seconds):
    """将秒数格式化为可读时长，如「2分15秒」「1小时3分」。"""
    if seconds is None:
        return "—"
    try:
        total = max(0, int(round(float(seconds))))
    except (TypeError, ValueError):
        return "—"
    if total < 60:
        return f"{total}秒"
    m, s = divmod(total, 60)
    if m < 60:
        return f"{m}分{s}秒" if s else f"{m}分"
    h, m = divmod(m, 60)
    if m:
        return f"{h}小时{m}分{s}秒" if s else f"{h}小时{m}分"
    return f"{h}小时{s}秒" if s else f"{h}小时"


def summarize_qc_timing(timing):
    """将 run_qc_batch 产出的 timing 汇总为可展示字段。"""
    if not timing:
        return {}

    wall = float(timing.get("wall_seconds") or 0)
    conc = int(timing.get("concurrency") or 1)
    light = timing.get("light") or {}
    full = timing.get("full") or {}
    skip = timing.get("skip") or {}

    light_n = int(light.get("count") or 0)
    full_n = int(full.get("count") or 0)
    skip_n = int(skip.get("count") or 0)
    light_sum = float(light.get("total_seconds") or 0)
    full_sum = float(full.get("total_seconds") or 0)
    skip_sum = float(skip.get("total_seconds") or 0)
    api_n = light_n + full_n
    total_task_sum = light_sum + full_sum + skip_sum

    def _avg(sum_s, n):
        return round(sum_s / n, 2) if n else None

    def _parallel_avg(tier_sum, tier_n):
        """按任务耗时占比，将墙钟时间分摊到该档，再除以通数。"""
        if not tier_n:
            return None
        if total_task_sum > 0:
            return round(wall * (tier_sum / total_task_sum) / tier_n, 2)
        return round(wall / api_n, 2) if api_n else None

    started = timing.get("started_at")
    ended = timing.get("ended_at")
    started_s = started.strftime("%Y-%m-%d %H:%M:%S") if started else ""
    ended_s = ended.strftime("%Y-%m-%d %H:%M:%S") if ended else ""

    return {
        "开始时间": started_s,
        "结束时间": ended_s,
        "总耗时秒": round(wall, 2),
        "总耗时文本": format_duration(wall),
        "并发数": conc,
        "轻量数量": light_n,
        "完整数量": full_n,
        "跳过数量": skip_n,
        "分析客户数": api_n,
        "分析客户均价秒": round(wall / api_n, 2) if api_n else None,
        "分析客户均价文本": format_duration(wall / api_n) if api_n else "—",
        "轻量单通耗时秒": _avg(light_sum, light_n),
        "轻量单通耗时文本": format_duration(_avg(light_sum, light_n)),
        "轻量并行折算秒": _parallel_avg(light_sum, light_n),
        "轻量并行折算文本": format_duration(_parallel_avg(light_sum, light_n)),
        "完整单通耗时秒": _avg(full_sum, full_n),
        "完整单通耗时文本": format_duration(_avg(full_sum, full_n)),
        "完整并行折算秒": _parallel_avg(full_sum, full_n),
        "完整并行折算文本": format_duration(_parallel_avg(full_sum, full_n)),
    }


def _qc_tier_wave_order():
    """质检 API 调用档位顺序：skip → full → light，同档连续跑利于 DeepSeek 上下文缓存命中。"""
    return ("skip", "full", "light")


def _group_session_indices_by_tier(sessions):
    """按档位分组会话，保留原始下标以便结果回填。"""
    buckets = {t: [] for t in _qc_tier_wave_order()}
    for idx, s in enumerate(sessions):
        cust, _ = count_roles(s.get("对话", ""))
        tier = classify_tier(cust)
        buckets[tier].append((idx, s))
    return [(tier, buckets[tier]) for tier in _qc_tier_wave_order() if buckets[tier]]


def _run_parallel_with_cache_warmup(items, conc, run_one, cache_warmup=True):
    """先串行跑 1 条预热 DeepSeek 上下文缓存，再满并发跑其余。

    items: 可迭代的一批任务入参；run_one(item) 返回任意结果。
    返回与 items 顺序一致的结果列表。
    """
    items = list(items)
    if not items:
        return []

    results = []
    if cache_warmup and len(items) > 1:
        results.append(run_one(items[0]))
        pending = items[1:]
    else:
        pending = items

    if not pending:
        return results

    workers = 1 if len(pending) == 1 else max(1, int(conc))
    with ThreadPoolExecutor(max_workers=workers) as ex:
        futs = [ex.submit(run_one, it) for it in pending]
        for fut in as_completed(futs):
            results.append(fut.result())
    return results


def _run_qc_tier_wave(cfg, items, conc, tier_stats, *, cache_warmup=False):
    """并发处理同一档位的一批会话，返回 {idx: (row, tier, elapsed)}。"""
    if not items:
        return {}

    def task(entry):
        idx, s = entry
        t0 = time.perf_counter()
        row, tier = qc_one_session(cfg, s)
        return idx, row, tier, time.perf_counter() - t0

    def _record(result):
        idx, row, tier, elapsed = result
        bucket = tier_stats.get(tier)
        if bucket is not None:
            bucket["count"] += 1
            bucket["total_seconds"] += elapsed
        return idx, (row, tier, elapsed)

    out = {}
    for result in _run_parallel_with_cache_warmup(items, conc, task, cache_warmup=cache_warmup):
        idx, packed = _record(result)
        out[idx] = packed
    return out


def run_qc_batch(sessions, cfg, on_progress=None, timing_out=None):
    """并发质检多通会话。按档位分批（skip → full → light）调度，提高 API 缓存命中。

    on_progress(done, total, row, tier) 可选回调。
    timing_out: 若传入 dict，会写入耗时统计原始数据（供 summarize_qc_timing 使用）。
    """
    results = [None] * len(sessions)
    done = 0
    total = len(sessions)
    conc = max(1, int(cfg.get("concurrency", DEFAULT_CONCURRENCY)))

    tier_stats = {
        "skip": {"count": 0, "total_seconds": 0.0},
        "light": {"count": 0, "total_seconds": 0.0},
        "full": {"count": 0, "total_seconds": 0.0},
    }
    started_at = datetime.now()
    batch_t0 = time.perf_counter()

    for wave_tier, items in _group_session_indices_by_tier(sessions):
        wave_out = _run_qc_tier_wave(
            cfg, items, conc, tier_stats,
            cache_warmup=(wave_tier != "skip"),
        )
        for idx in sorted(wave_out.keys()):
            row, tier, _elapsed = wave_out[idx]
            results[idx] = row
            done += 1
            if on_progress:
                on_progress(done, total, row, tier)

    wall_seconds = time.perf_counter() - batch_t0
    ended_at = datetime.now()
    raw_timing = {
        "started_at": started_at,
        "ended_at": ended_at,
        "wall_seconds": wall_seconds,
        "concurrency": conc,
        **tier_stats,
    }
    if timing_out is not None:
        timing_out.clear()
        timing_out.update(raw_timing)

    return results


# 已废弃字段：若旧 prompt/缓存返回这些字段，导出前一律剔除（评分体系 + 已删除的「做得好」列）。
_DROPPED_SCORE_COLUMNS = frozenset({
    "总分", "分_流程", "分_心理应答", "分_异议", "分_激活", "分_逼单",
    "做得好",
})


def build_report_dataframe(results, columns=None):
    cols = columns or REPORT_COLUMNS
    out_df = pd.DataFrame(results)
    out_df = out_df.drop(columns=[c for c in _DROPPED_SCORE_COLUMNS if c in out_df.columns], errors="ignore")
    for c in cols:
        if c not in out_df.columns:
            out_df[c] = ""
    extra = [c for c in out_df.columns if c not in cols]
    return out_df[cols + extra]


def build_low_intent_dataframe(results):
    return build_report_dataframe(results, LOW_INTENT_REPORT_COLUMNS)


def build_deep_review_dataframe(results):
    return build_report_dataframe(results or [], DEEP_REVIEW_REPORT_COLUMNS)


REPORT_SHEET_QUALITY = "优质客户分析"
REPORT_SHEET_LOW_INTENT = "低意向客户"


def _is_low_intent_row(row):
    """客户发言不足 3 条，跳过 API 质检的低意向会话。"""
    tier = str(row.get("质检档位", "")).strip()
    if tier == TIER_LABEL["skip"]:
        return True
    try:
        return int(row.get("客户发言数", 0)) < TIER_SKIP_MAX
    except (TypeError, ValueError):
        return False


def split_report_results(results):
    """按客户发言数拆分为优质客户与低意向客户两组。"""
    quality, low_intent = _split_results(results)
    return build_report_dataframe(quality), build_low_intent_dataframe(low_intent)


def _split_results(results):
    quality, low_intent = [], []
    for row in results:
        (low_intent if _is_low_intent_row(row) else quality).append(row)
    return quality, low_intent


def _highlight_quality_sheet(writer, quality_rows, report_columns):
    """优质客户分析：仅将有问题的单元格标红。"""
    from openpyxl.styles import PatternFill

    ws = writer.sheets.get(REPORT_SHEET_QUALITY)
    if ws is None or not quality_rows:
        return

    red_fill = PatternFill(fill_type="solid", start_color="FFFFC7CE", end_color="FFFFC7CE")
    col_index = {name: idx + 1 for idx, name in enumerate(report_columns)}

    for row_idx, row in enumerate(quality_rows, start=2):
        for col_name in problem_columns_for_row(row):
            col_idx = col_index.get(col_name)
            if col_idx:
                ws.cell(row=row_idx, column=col_idx).fill = red_fill


def _highlight_low_intent_sheet(writer, low_intent_rows, report_columns):
    """低意向客户：统计时段内客服首次回复超 10 分钟时标红对应单元格。"""
    from openpyxl.styles import PatternFill

    ws = writer.sheets.get(REPORT_SHEET_LOW_INTENT)
    if ws is None or not low_intent_rows:
        return

    red_fill = PatternFill(fill_type="solid", start_color="FFFFC7CE", end_color="FFFFC7CE")
    col_index = {name: idx + 1 for idx, name in enumerate(report_columns)}

    for row_idx, row in enumerate(low_intent_rows, start=2):
        for col_name in problem_columns_for_low_intent_row(row):
            col_idx = col_index.get(col_name)
            if col_idx:
                ws.cell(row=row_idx, column=col_idx).fill = red_fill


def write_report_excel(results, path_or_buffer, deep_review_rows=None):
    """导出 Excel：优质客户分析 + 低意向客户 + 可选深度复盘。"""
    quality_rows, low_intent_rows = _split_results(results)
    df_quality = build_report_dataframe(quality_rows)
    df_low = build_low_intent_dataframe(low_intent_rows)
    deep_rows = [r for r in (deep_review_rows or []) if r]
    df_deep = build_deep_review_dataframe(deep_rows) if deep_rows else None
    with pd.ExcelWriter(path_or_buffer, engine="openpyxl") as writer:
        df_quality.to_excel(writer, sheet_name=REPORT_SHEET_QUALITY, index=False)
        df_low.to_excel(writer, sheet_name=REPORT_SHEET_LOW_INTENT, index=False)
        if df_deep is not None and len(df_deep):
            df_deep.to_excel(writer, sheet_name=REPORT_SHEET_DEEP_REVIEW, index=False)
        _highlight_quality_sheet(writer, quality_rows, list(df_quality.columns))
        _highlight_low_intent_sheet(writer, low_intent_rows, list(df_low.columns))


# ============================================================
# 聊天进度对比（同一天多份质检报告逐段比对）
# ============================================================
_REPORT_NAME_TS_RE = re.compile(r"(\d{8})[_-](\d{6})")
PROGRESS_COMPARE_SOURCE_SHEETS = (REPORT_SHEET_QUALITY, REPORT_SHEET_LOW_INTENT)


def parse_report_timestamp_from_name(name):
    """从报告文件名解析快照时间，如 质检报告_自定义_20260615_105523.xlsx。"""
    m = _REPORT_NAME_TS_RE.search(str(name or ""))
    if not m:
        return None
    try:
        return datetime.strptime(m.group(1) + m.group(2), "%Y%m%d%H%M%S")
    except ValueError:
        return None


def _customer_key(row):
    """客户匹配主键：优先会话ID，否则联系人。"""
    sid = str(row.get("会话ID", "") or "").strip()
    if sid:
        return ("id", sid)
    name = str(row.get("联系人", "") or "").strip()
    if name:
        return ("name", name)
    return None


def load_report_customers(sheets):
    """从单份报告的工作表字典中合并出 {客户key: 字段dict}。

    sheets: {工作表名: DataFrame}。合并「优质客户分析」「低意向客户」两表，
    并按会话ID补充「深度复盘」的推荐下一步动作。
    """
    customers = {}
    for sheet_name in PROGRESS_COMPARE_SOURCE_SHEETS:
        df = sheets.get(sheet_name)
        if df is None:
            continue
        for record in df.fillna("").to_dict("records"):
            key = _customer_key(record)
            if key is None:
                continue
            customers.setdefault(key, dict(record))

    deep_df = sheets.get(REPORT_SHEET_DEEP_REVIEW)
    if deep_df is not None:
        for record in deep_df.fillna("").to_dict("records"):
            key = _customer_key(record)
            if key is None or key not in customers:
                continue
            rec = customers[key]
            for col in ("推荐下一步动作", "跟进话术建议", "卡点诊断"):
                if record.get(col) and not rec.get(col):
                    rec[col] = record[col]
    return customers


def _median_last_customer_ts(customers):
    """以各客户原始对话末条消息时间的中位数作为报告快照时间兜底。"""
    stamps = []
    for row in customers.values():
        ts = _last_customer_message_timestamp(str(row.get("原始对话", "") or ""))
        if ts is not None:
            stamps.append(ts)
    if not stamps:
        return None
    stamps.sort()
    return stamps[len(stamps) // 2]


def build_report_meta(name, sheets):
    """构造单份报告的元信息：名称、客户表、快照时间与来源。"""
    customers = load_report_customers(sheets)
    ts = parse_report_timestamp_from_name(name)
    source = "文件名"
    if ts is None:
        ts = _median_last_customer_ts(customers)
        source = "末条消息中位数" if ts is not None else "未知"
    return {
        "报告名": name,
        "客户": customers,
        "快照时间": ts,
        "快照时间来源": source,
    }


def sort_report_metas(metas):
    """按快照时间从早到晚排序；无时间者排在最后，保持原始相对顺序。"""
    def _key(item):
        idx, meta = item
        ts = meta.get("快照时间")
        if ts is not None:
            return (0, ts, idx)
        return (1, datetime.max, idx)

    ordered = sorted(enumerate(metas), key=_key)
    return [meta for _, meta in ordered]


def _max_line_timestamp(text):
    """对话中最晚一条可解析的发言时间。"""
    latest = None
    for line in str(text or "").splitlines():
        ts = _parse_line_timestamp(line)
        if ts is not None and (latest is None or ts > latest):
            latest = ts
    return latest


def _line_suffix_delta(earlier_text, later_text):
    """无时间戳时的兜底：去掉两段对话的公共前缀行，返回较晚独有的后续行。"""
    e_lines = [l for l in str(earlier_text or "").splitlines() if l.strip()]
    l_lines = [l for l in str(later_text or "").splitlines() if l.strip()]
    i = 0
    while i < len(e_lines) and i < len(l_lines) and e_lines[i].strip() == l_lines[i].strip():
        i += 1
    return "\n".join(l_lines[i:])


def extract_dialogue_delta(earlier_text, later_text):
    """提取较晚报告相对较早报告新增的对话片段。

    优先按时间戳：取较早对话中最晚一条发言时间为界，返回较晚对话里时间更晚的行；
    较早对话无可解析时间时，退化为去公共前缀的文本后缀 diff。
    """
    cutoff = _max_line_timestamp(earlier_text)
    if cutoff is not None:
        kept = []
        for line in str(later_text or "").splitlines():
            if not line.strip():
                continue
            ts = _parse_line_timestamp(line)
            if ts is not None and ts > cutoff:
                kept.append(line)
        return "\n".join(kept)
    return _line_suffix_delta(earlier_text, later_text)


def build_progress_compare_pairs(sorted_metas):
    """相邻两份报告间，为同时出现的客户生成逐段对比单元。"""
    pairs = []
    for seg_idx in range(len(sorted_metas) - 1):
        earlier = sorted_metas[seg_idx]
        later = sorted_metas[seg_idx + 1]
        seg_label = f"{seg_idx + 1}→{seg_idx + 2}"
        for key, later_row in later["客户"].items():
            earlier_row = earlier["客户"].get(key)
            if earlier_row is None:
                continue
            delta = extract_dialogue_delta(
                earlier_row.get("原始对话", ""),
                later_row.get("原始对话", ""),
            )
            pairs.append({
                "联系人": later_row.get("联系人", "") or earlier_row.get("联系人", ""),
                "会话ID": later_row.get("会话ID", "") or earlier_row.get("会话ID", ""),
                "接待成员": later_row.get("接待成员", "") or earlier_row.get("接待成员", ""),
                "渠道": later_row.get("渠道", "") or earlier_row.get("渠道", ""),
                "对比段": seg_label,
                "较早报告": earlier["报告名"],
                "较晚报告": later["报告名"],
                "较早": earlier_row,
                "较晚": later_row,
                "新增对话": delta,
            })
    return pairs


def _format_progress_snapshot(label, name, row):
    fields = [
        f"结果标签：{row.get('结果标签', '')}",
        f"客户阶段：{row.get('客户阶段', '')}",
        f"是否合格：{row.get('是否合格', '')}",
        f"风险等级：{row.get('风险等级', '')}",
        f"客户主要顾虑：{row.get('客户主要顾虑', '')}",
        f"主要问题：{row.get('主要问题', '')}",
        f"改进建议：{row.get('改进建议', '')}",
        f"下一步跟进动作：{row.get('下一步跟进动作', '')}",
    ]
    deep = row.get("推荐下一步动作")
    if deep:
        fields.append(f"深度复盘推荐动作：{deep}")
    fields.append(f"客户发言数：{row.get('客户发言数', '')} · 客服发言数：{row.get('客服发言数', '')}")
    body = "\n".join(str(f) for f in fields)
    return f"【{label}】（报告：{name}）\n{body}"


def build_progress_compare_payload(pair):
    """拼出供大模型阅读的完整对比载荷（较早快照 + 新增对话 + 较晚快照）。"""
    delta = pair.get("新增对话") or "（未提取到明显新增对话，请结合两次快照字段判断）"
    parts = [
        _format_progress_snapshot("较早快照", pair.get("较早报告", ""), pair.get("较早", {})),
        "",
        "【两次快照之间的新增对话】",
        delta,
        "",
        _format_progress_snapshot("较晚快照", pair.get("较晚报告", ""), pair.get("较晚", {})),
    ]
    return "\n".join(parts)


def compute_progress_touch_counts(pair):
    """进度对比：按 10 分钟规则统计较早/新增/今日销售触达次数。"""
    earlier_text = str((pair.get("较早") or {}).get("原始对话", "") or "")
    later_text = str((pair.get("较晚") or {}).get("原始对话", "") or "")
    delta = str(pair.get("新增对话", "") or "")
    return {
        "较早销售触达次数": count_service_touch_bursts(earlier_text),
        "新增销售触达次数": count_service_touch_bursts(delta),
        "今日销售触达次数": count_service_touch_bursts(later_text),
    }


def compare_progress_one(cfg, pair):
    """对单个客户的相邻两份快照做进度对比，异常不外抛。"""
    earlier = pair.get("较早", {})
    later = pair.get("较晚", {})
    delta = pair.get("新增对话", "")
    row = {
        "联系人": pair.get("联系人", ""),
        "会话ID": pair.get("会话ID", ""),
        "接待成员": pair.get("接待成员", ""),
        "渠道": pair.get("渠道", ""),
        "对比段": pair.get("对比段", ""),
        "较早报告": pair.get("较早报告", ""),
        "较晚报告": pair.get("较晚报告", ""),
        "较早结果标签": earlier.get("结果标签", ""),
        "较晚结果标签": later.get("结果标签", ""),
        "新增对话": delta,
        **compute_progress_touch_counts(pair),
        "新增客户发言数": count_valid_customer_segments(delta),
        "新增客服发言数": count_valid_agent_segments(delta),
    }
    try:
        res = call_llm(cfg, build_progress_compare_payload(pair), tier="progress_compare")
        for k, v in res.items():
            val = v if not isinstance(v, (dict, list)) else json.dumps(v, ensure_ascii=False)
            if isinstance(val, str):
                val = redact_secrets(val, cfg)
            row[k] = val
    except Exception as e:
        row["_错误"] = redact_secrets(str(e), cfg)
    return row


def run_progress_compare_batch(pairs, cfg, on_progress=None, timing_out=None):
    """并发执行进度对比。on_progress(done, total, row) 可选回调。"""
    results = [None] * len(pairs)
    if not pairs:
        if timing_out is not None:
            timing_out.clear()
        return results

    done = 0
    total = len(pairs)
    conc = max(1, int(cfg.get("concurrency", DEFAULT_CONCURRENCY)))
    started_at = datetime.now()
    batch_t0 = time.perf_counter()
    total_seconds = 0.0

    def task(entry):
        i, pair = entry
        t0 = time.perf_counter()
        row = compare_progress_one(cfg, pair)
        return i, row, time.perf_counter() - t0

    indexed = list(enumerate(pairs))
    for i, row, elapsed in _run_parallel_with_cache_warmup(indexed, conc, task, cache_warmup=True):
        results[i] = row
        total_seconds += elapsed
        done += 1
        if on_progress:
            on_progress(done, total, row)

    if timing_out is not None:
        timing_out.clear()
        timing_out.update({
            "started_at": started_at,
            "ended_at": datetime.now(),
            "wall_seconds": time.perf_counter() - batch_t0,
            "concurrency": conc,
            "count": total,
            "total_seconds": total_seconds,
        })
    return results


def build_progress_compare_dataframe(rows):
    return build_report_dataframe(rows or [], PROGRESS_COMPARE_REPORT_COLUMNS)


def _progress_compare_problem_columns(row):
    """进度对比中需要标红的列：失败、高风险、问题未补救、未按建议执行。"""
    cols = []
    if row.get("_错误"):
        cols.append("_错误")
    if _is_high_risk_level(row.get("风险等级")):
        cols.append("风险等级")
    remedy = str(row.get("问题是否补救", "")).strip()
    if remedy in ("否", "未补救", "没有"):
        cols.append("问题是否补救")
    follow = str(row.get("是否按建议执行", "")).strip()
    if follow in ("否", "未执行", "没有"):
        cols.append("是否按建议执行")
    return cols


def _highlight_progress_compare_sheet(writer, rows, report_columns):
    from openpyxl.styles import PatternFill

    ws = writer.sheets.get(REPORT_SHEET_PROGRESS_COMPARE)
    if ws is None or not rows:
        return
    red_fill = PatternFill(fill_type="solid", start_color="FFFFC7CE", end_color="FFFFC7CE")
    col_index = {name: idx + 1 for idx, name in enumerate(report_columns)}
    for row_idx, row in enumerate(rows, start=2):
        for col_name in _progress_compare_problem_columns(row):
            col_idx = col_index.get(col_name)
            if col_idx:
                ws.cell(row=row_idx, column=col_idx).fill = red_fill


def write_progress_compare_excel(rows, path_or_buffer):
    """导出进度对比报告（单工作表「进度对比」）。"""
    clean = [r for r in (rows or []) if r]
    df = build_progress_compare_dataframe(clean)
    with pd.ExcelWriter(path_or_buffer, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=REPORT_SHEET_PROGRESS_COMPARE, index=False)
        _highlight_progress_compare_sheet(writer, clean, list(df.columns))


def compute_progress_compare_stats(rows):
    """进度对比结果汇总。"""
    rows = [r for r in (rows or []) if r]
    n_total = len(rows)
    n_high = sum(1 for r in rows if _is_high_risk_level(r.get("风险等级")) or r.get("_错误"))
    n_not_remedied = sum(
        1 for r in rows if str(r.get("问题是否补救", "")).strip() in ("否", "未补救", "没有")
    )
    n_not_followed = sum(
        1 for r in rows if str(r.get("是否按建议执行", "")).strip() in ("否", "未执行", "没有")
    )
    segments = sorted({str(r.get("对比段", "")).strip() for r in rows if r.get("对比段")})
    return {
        "对比客户对数": n_total,
        "高风险数量": n_high,
        "问题未补救数量": n_not_remedied,
        "未按建议执行数量": n_not_followed,
        "对比段": segments,
    }


def compute_stats(results):
    """汇总统计。"""
    n_skip = n_light = n_full = n_risk = 0
    for row in results:
        tier_key = {v: k for k, v in TIER_LABEL.items()}.get(row.get("质检档位"), "")
        if tier_key == "skip":
            n_skip += 1
        elif tier_key == "light":
            n_light += 1
        elif tier_key == "full":
            n_full += 1
        if is_high_risk(row):
            n_risk += 1
    return {
        "总会话数": len(results),
        "跳过数量": n_skip,
        "轻量质检数量": n_light,
        "完整质检数量": n_full,
        "高风险数量": n_risk,
    }
